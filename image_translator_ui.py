"""
图片翻译 UI 功能区 — Streamlit Tab3 业务逻辑层。

提供文件校验、翻译编排、下载生成等纯数据操作，
供 app.py 的 Streamlit UI 层调用。
"""

import os
import tempfile
from dataclasses import dataclass, field
from pathlib import Path


@dataclass
class ImageUploadResult:
    """图片翻译上传校验结果。

    Attributes:
        is_valid: 校验是否通过。
        error: 失败时的中文错误信息。
        products: 成功时返回的产品列表。
        count: ASIN 数量。
        input_type: "crawler_output"（4列）或 "translation_output"（12列）。
    """
    is_valid: bool = False
    error: str = ""
    products: list[dict] = field(default_factory=list)
    count: int = 0
    input_type: str = ""


def validate_image_upload(filepath: str | Path) -> ImageUploadResult:
    """校验上传的 Excel 文件。

    支持两种格式：
    - 爬虫 4 列输出（asin, 图片url, 标题, 详情）
    - 翻译后 12 列输出（含核心流量词, 俄语标题, 俄语详情 等）

    Args:
        filepath: 上传的 Excel 文件路径。

    Returns:
        ImageUploadResult: 校验结果，含产品列表和格式类型。
    """
    filepath = Path(filepath)
    filename = filepath.name

    if not filename.lower().endswith(".xlsx"):
        return ImageUploadResult(
            is_valid=False,
            error="仅支持 .xlsx 格式的 Excel 文件，请重新上传。",
            count=0,
        )

    import openpyxl

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
    except Exception as e:
        return ImageUploadResult(
            is_valid=False,
            error=f"无法打开 Excel 文件: {e}",
            count=0,
        )

    # 读取表头
    headers = [cell.value for cell in ws[1]]
    if headers is None:
        wb.close()
        return ImageUploadResult(
            is_valid=False,
            error="Excel 文件为空或没有表头行。",
            count=0,
        )

    header_set = {str(h).strip() if h is not None else "" for h in headers}

    # 检查必要列
    if "图片url" not in header_set:
        wb.close()
        return ImageUploadResult(
            is_valid=False,
            error="缺少必要列「图片url」。请确保 Excel 包含图片url 列。",
            count=0,
        )

    if "asin" not in header_set:
        wb.close()
        return ImageUploadResult(
            is_valid=False,
            error="缺少必要列「asin」。请确保 Excel 包含 asin 列。",
            count=0,
        )

    # 确定输入类型
    if "俄语标题" in header_set and "核心流量词" in header_set:
        input_type = "translation_output"
    else:
        input_type = "crawler_output"

    # 构建列索引映射
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = str(h).strip() if h is not None else ""
        col_map[key] = i

    # 读取数据行
    products: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        asin = _safe_cell(row, col_map.get("asin"))
        image_urls = _safe_cell(row, col_map.get("图片url"))

        if not asin or not image_urls:
            continue

        product = {"asin": asin, "图片url": image_urls}

        # 构建产品上下文信息库（用于图片翻译时传给 AI 做上下文感知翻译）
        product_context: dict[str, str] = {}

        # 始终读取英文原始列（如有）
        for col_name in ("标题", "详情"):
            if col_name in col_map:
                val = _safe_cell(row, col_map[col_name])
                if val:
                    product[col_name] = val
                    product_context[col_name] = val

        # 如果是翻译后 12 列格式，读取俄语列作为上下文
        if input_type == "translation_output":
            for col_name in ("俄语标题", "核心流量词", "俄语详情"):
                if col_name in col_map:
                    val = _safe_cell(row, col_map[col_name])
                    if val:
                        product_context[col_name] = val

        # 始终挂载 product_context（即使为空字典）
        product["product_context"] = product_context

        products.append(product)

    wb.close()

    return ImageUploadResult(
        is_valid=True,
        error="",
        products=products,
        count=len(products),
        input_type=input_type,
    )


def _safe_cell(row: tuple, col_idx: int | None) -> str:
    """安全读取单元格值。"""
    if col_idx is None:
        return ""
    if col_idx >= len(row):
        return ""
    val = row[col_idx]
    if val is None:
        return ""
    return str(val).strip()


def start_background_translation(
    products: list[dict],
    font_config=None,
) -> None:
    """启动后台图片翻译 Worker。

    Worker 在独立 asyncio task 中运行，不阻塞 UI。

    Args:
        products: 产品列表。
        font_config: 字体配置。
    """
    from worker import WorkerManager

    manager = WorkerManager()
    manager.start(products=products, font_config=font_config)


def get_worker_status():
    """读取当前 Worker 状态。

    Returns:
        WorkerStatus: 当前进度和状态信息。
    """
    from worker import WorkerManager

    manager = WorkerManager()
    return manager.get_status()


def pause_worker() -> None:
    """暂停后台 Worker。"""
    from worker import WorkerManager

    WorkerManager().pause()


def resume_worker() -> None:
    """恢复后台 Worker。"""
    from worker import WorkerManager

    WorkerManager().resume()


def run_image_translation(
    products: list[dict],
    font_config=None,
):
    """同步运行图片翻译（阻塞直到完成），返回 BatchImageResult。

    用于 Streamlit UI 等需要等待结果的场景。

    Args:
        products: 产品列表。
        font_config: 字体配置。

    Returns:
        BatchImageResult。
    """
    from worker import WorkerManager

    manager = WorkerManager()
    return manager.run_sync(products=products, font_config=font_config)


def generate_output_excel(results) -> bytes:
    """生成翻译后的输出 Excel（图片 URL 已替换为 R2 URL）。

    Args:
        results: BatchImageResult 实例。

    Returns:
        xlsx 文件的二进制数据。
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active

    # 表头（与爬虫输出格式一致 + 新增 R2 URL 列）
    ws.append(["asin", "图片url", "r2图片url", "俄语图片状态"])

    for asin_result in results.results:
        asin = asin_result.asin
        for img in asin_result.images:
            ws.append([
                asin,
                img.original_url,
                img.r2_url,
                img.status,
            ])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb.save(tmp_path)
        wb.close()
        with open(tmp_path, "rb") as f:
            return f.read()
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def generate_report(results) -> bytes:
    """生成处理报告（CSV 格式，含每张图的状态和错误信息）。

    Args:
        results: BatchImageResult 实例。

    Returns:
        CSV 文件的二进制数据（UTF-8 BOM）。
    """
    import csv
    import io

    output = io.StringIO()
    writer = csv.writer(output)

    # 表头
    writer.writerow(["asin", "图片序号", "原始URL", "R2 URL", "状态", "错误信息", "原文", "译文"])

    for asin_result in results.results:
        asin = asin_result.asin
        for img in asin_result.images:
            writer.writerow([
                asin,
                img.index,
                img.original_url,
                img.r2_url,
                img.status,
                img.error,
                "; ".join(img.ocr_original_texts) if img.ocr_original_texts else "",
                "; ".join(img.translated_texts) if img.translated_texts else "",
            ])

    # UTF-8 BOM（Excel 兼容）
    return ("﻿" + output.getvalue()).encode("utf-8")


# ============================================================
# 单图重试功能
# ============================================================


def retry_single_image(
    image_url: str,
    asin: str,
    index: int,
    font_config,
    product_context: dict[str, str] | None = None,
) -> "ImageResult":
    """同步重试单张图片翻译。

    内部调用 translate_single_image，通过 asyncio.run() 桥接异步管线。
    用于 Streamlit UI 中用户点击"重新翻译"按钮后的即时响应。

    Args:
        image_url: 原始图片 URL。
        asin: 产品 ASIN。
        index: 图片序号（0-based）。
        font_config: 字体配置。
        product_context: 产品数据信息库（可选）。

    Returns:
        ImageResult: 新的处理结果。
    """
    import asyncio

    from image_translator import translate_single_image

    return asyncio.run(
        translate_single_image(
            image_url=image_url,
            asin=asin,
            index=index,
            font_config=font_config,
            product_context=product_context,
        )
    )


def _find_product_context(
    products: list[dict], asin: str
) -> dict[str, str] | None:
    """从产品列表中按 ASIN 查找 product_context。

    Args:
        products: validate_image_upload 产出的产品列表。
        asin: 目标 ASIN。

    Returns:
        product_context 字典，若未找到则返回 None。
    """
    for product in products:
        if product.get("asin") == asin:
            return product.get("product_context", None)
    return None


def enrich_product_context_from_db(
    products: list[dict],
    db_path: str = "products.db",
) -> list[dict]:
    """从数据库读取 Phase 1/2 产物，丰富产品上下文信息库。

    数据库中的字段优先级高于 Excel 中的同名字段。
    Phase 2 未完成时仅提供 Phase 1 产物（英文属性）；
    Phase 2 完成后额外提供俄语文案作为翻译参考。

    Args:
        products: validate_image_upload 产出的产品列表（原地修改）。
        db_path: SQLite 数据库文件路径。

    Returns:
        原地修改后的 products 列表（方便链式调用）。
    """
    import sqlite3

    from db import get_product, get_translation, init_db

    if not os.path.isfile(db_path):
        return products

    db = sqlite3.connect(db_path)
    init_db(db)

    try:
        for product in products:
            asin = product.get("asin", "")
            if not asin:
                continue

            product_info = get_product(db, asin)
            translation = get_translation(db, asin)

            ctx: dict[str, str] = product.get("product_context", {})

            # Phase 1 产物：英文属性（DB 优先，Excel 回退）
            if product_info:
                if product_info.get("title"):
                    ctx["标题"] = product_info["title"]
                if product_info.get("details"):
                    ctx["详情"] = product_info["details"]
                if product_info.get("category"):
                    ctx["product_category"] = product_info["category"]
                if product_info.get("material"):
                    ctx["material"] = product_info["material"]
                features = product_info.get("features", [])
                if features:
                    ctx["features"] = ", ".join(features[:5]) if isinstance(features, list) else str(features)
                if product_info.get("target_audience"):
                    ctx["target_audience"] = product_info["target_audience"]
                if product_info.get("brand"):
                    ctx["brand"] = product_info["brand"]

            # Phase 2 产物：俄语文案（DB 优先，Excel 回退）
            if translation:
                if translation.get("russian_title"):
                    ctx["俄语标题"] = translation["russian_title"]
                if translation.get("core_keywords"):
                    ctx["核心流量词"] = translation["core_keywords"]
                if translation.get("russian_description"):
                    ctx["俄语详情"] = translation["russian_description"]

            # Excel 回退：DB 无数据时保留 Excel 中已有的值
            # （ctx 中已有的 Excel 值不会被空 DB 字段覆盖）

            product["product_context"] = ctx

    finally:
        db.close()

    return products


def _replace_and_recount(
    batch_result: "BatchImageResult",
    asin: str,
    image_index: int,
    new_result: "ImageResult",
) -> None:
    """在批量结果中替换单张图片结果并重算所有计数字段。

    原地修改 batch_result。

    Args:
        batch_result: 当前的批量处理结果。
        asin: 目标 ASIN。
        image_index: 图片序号（ImageResult.index）。
        new_result: 新的处理结果（retry 产出的 ImageResult）。
    """
    # 1. 找到 AsinImageResult，替换 ImageResult，累加 retry_count
    for asin_result in batch_result.results:
        if asin_result.asin == asin:
            for i, img in enumerate(asin_result.images):
                if img.index == image_index:
                    new_result.retry_count = img.retry_count + 1
                    asin_result.images[i] = new_result
                    break

            # 2. 重算 AsinImageResult 计数
            asin_result.success_count = sum(
                1 for r in asin_result.images if r.status == "ok"
            )
            asin_result.error_count = sum(
                1 for r in asin_result.images if r.status == "error"
            )
            asin_result.skipped_count = sum(
                1 for r in asin_result.images if r.status == "skipped"
            )
            asin_result.video_count = sum(
                1 for r in asin_result.images if r.status == "video"
            )
            break

    # 3. 重算 BatchImageResult 顶层计数
    batch_result.success_images = sum(
        r.success_count for r in batch_result.results
    )
    batch_result.error_images = sum(
        r.error_count for r in batch_result.results
    )
    batch_result.skipped_images = sum(
        r.skipped_count for r in batch_result.results
    )
    batch_result.video_images = sum(
        r.video_count for r in batch_result.results
    )
    batch_result.total_images = sum(
        len(r.images) for r in batch_result.results
    )


# ============================================================
# 合并表格：将处理后的 R2 URL 写回原始 Excel
# ============================================================


def generate_merged_excel(
    original_file_bytes: bytes,
    url_mapping: dict[str, str],
) -> bytes:
    """将原始 Excel 中的 图片url 列替换为处理后的 R2 URL。

    保留原始表格的全部行列结构，仅替换 图片url 列的内容。
    新增两列：
      - R2图片url: 处理后的 R2 URL（便于对比）
      - 翻译状态: ✅/❌ 标记每张图片的处理结果

    Args:
        original_file_bytes: 原始上传 Excel 文件的字节数据。
        url_mapping: {原始图片URL: R2 URL} 映射。未在映射中的 URL 保持原样。

    Returns:
        合并后的 xlsx 文件字节数据，可直接用于 Streamlit download_button。
    """
    import re
    import tempfile

    from io import BytesIO

    import openpyxl

    # 读取原始 Excel
    wb = openpyxl.load_workbook(BytesIO(original_file_bytes))
    ws = wb.active

    # 读取表头
    headers = [cell.value for cell in ws[1]]
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = str(h).strip() if h is not None else ""
        col_map[key] = i

    image_col = col_map.get("图片url")
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # 新表头：原始列 + R2图片url + 翻译状态
    new_headers = list(headers)
    has_r2_col = "R2图片url" in col_map
    has_status_col = "翻译状态" in col_map
    if not has_r2_col:
        new_headers.append("R2图片url")
    if not has_status_col:
        new_headers.append("翻译状态")
    new_ws.append(new_headers)

    # 遍历数据行
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        new_row = list(row)

        # 替换图片url 列
        r2_urls: list[str] = []
        statuses: list[str] = []

        if image_col is not None and image_col < len(row):
            original_urls = str(row[image_col] or "")
            urls = re.split(r'[;|]', original_urls)

            for u in urls:
                u = u.strip()
                if not u:
                    continue
                r2 = url_mapping.get(u, u)
                r2_urls.append(r2)
                statuses.append("✅" if r2 != u else "❌")

            # 替换原始列
            # 保留原始数据但写入 R2 URLs（使用 | 分隔与原始格式一致）
            new_row[image_col] = "; ".join(r2_urls) if r2_urls else str(row[image_col] or "")

        # 补充 R2 列和状态列
        new_row.append("; ".join(r2_urls) if r2_urls else "")
        new_row.append("; ".join(statuses) if statuses else "")

        # 确保长度与新表头一致
        while len(new_row) < len(new_headers):
            new_row.append("")

        new_ws.append(new_row)

    wb.close()

    # 写入临时文件再读取为 bytes
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        new_wb.save(tmp_path)
        new_wb.close()
        with open(tmp_path, "rb") as f:
            return f.read()
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
