"""
Excel 读写模块 — 公共接口。

爬虫采集和文案翻译两个功能区共用此模块处理 Excel 文件。
"""

from pathlib import Path

import openpyxl


def read_asins_from_excel(filepath: str | Path) -> list[str]:
    """从 Excel 文件中读取 ASIN 列表。

    自动识别 ASIN 列（大小写不敏感），跳过空行，去除首尾空格。

    Args:
        filepath: Excel 文件路径。

    Returns:
        ASIN 字符串列表（保持原顺序，跳过空值）。

    Raises:
        ValueError: 找不到 ASIN 列时抛出。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # 读取表头行
    headers = [cell.value for cell in ws[1]]
    if headers is None:
        raise ValueError("Excel 文件为空或没有表头行")

    # 查找 ASIN 列索引（大小写不敏感）
    asin_col_idx: int | None = None
    for i, h in enumerate(headers):
        if h is not None and str(h).strip().lower() == "asin":
            asin_col_idx = i
            break

    if asin_col_idx is None:
        raise ValueError(
            f"找不到 ASIN 列。表头为: {headers}。"
            f"请确保 Excel 第一行包含 'asin' 列（大小写不敏感）。"
        )

    # 读取 ASIN 值（跳过空行和表头行）
    asins: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[asin_col_idx] if len(row) > asin_col_idx else None
        if val is not None:
            stripped = str(val).strip()
            if stripped:
                asins.append(stripped)

    wb.close()
    return asins


def write_products_to_excel(products: list[dict], filepath: str | Path) -> None:
    """将产品列表写入符合"爬虫表格（处理前）"格式的 Excel 文件。

    输出 4 列：asin、图片url、标题、详情。

    Args:
        products: 产品字典列表，每个字典需包含 asin, 图片url, 标题, 详情 键。
        filepath: 输出文件路径。
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # 写入表头（与"爬虫表格（处理前）"格式一致）
    ws.append(["asin", "图片url", "标题", "详情"])

    # 写入数据行
    for p in products:
        ws.append([
            p.get("asin", ""),
            p.get("图片url", ""),
            p.get("标题", ""),
            p.get("详情", ""),
        ])

    wb.save(filepath)
    wb.close()


# "爬虫表格（处理前）"必须包含的列
_PRE_COLUMNS = ["asin", "图片url", "标题", "详情"]


def read_products_from_excel(filepath: str | Path) -> list[dict]:
    """从"爬虫表格（处理前）"Excel 读取完整产品信息。

    Args:
        filepath: Excel 文件路径，需包含 asin, 图片url, 标题, 详情 四列。

    Returns:
        产品字典列表，每个字典包含 asin, 图片url, 标题, 详情 键。

    Raises:
        ValueError: 缺少必要列时抛出。
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    # 读取表头行
    headers = [cell.value for cell in ws[1]]
    if headers is None:
        raise ValueError("Excel 文件为空或没有表头行")

    # 检查必要列是否存在
    header_set = {str(h).strip() if h is not None else "" for h in headers}
    missing = [c for c in _PRE_COLUMNS if c not in header_set]
    if missing:
        raise ValueError(
            f"Excel 缺少必要列: {', '.join(missing)}。"
            f"需要包含: {', '.join(_PRE_COLUMNS)}"
        )

    # 构建列索引映射
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = str(h).strip() if h is not None else ""
        if key in _PRE_COLUMNS:
            col_map[key] = i

    # 读取数据行
    products: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue  # 跳过全空行
        product = {
            "asin": str(row[col_map["asin"]]).strip() if row[col_map["asin"]] is not None else "",
            "图片url": str(row[col_map["图片url"]]).strip() if row[col_map["图片url"]] is not None else "",
            "标题": str(row[col_map["标题"]]).strip() if row[col_map["标题"]] is not None else "",
            "详情": str(row[col_map["详情"]]).strip() if row[col_map["详情"]] is not None else "",
        }
        products.append(product)

    wb.close()
    return products


# "爬虫表格（处理后）" 12 列表头
_OUTPUT_COLUMNS = [
    "asin", "图片url", "标题", "详情",
    "核心流量词", "俄语标题", "俄语详情",
    "货源", "采购价", "", "", "商品类别",
]


def write_translated_products_to_excel(
    products: list[dict], filepath: str | Path
) -> None:
    """将翻译后的产品列表写入"爬虫表格（处理后）"格式的 Excel。

    输出 12 列，后 5 列（货源、采购价、空列、商品类别）为空。

    Args:
        products: 产品字典列表，需包含前 7 列的键。
        filepath: 输出文件路径。
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # 写入表头
    ws.append(_OUTPUT_COLUMNS)

    # 写入数据行
    for p in products:
        ws.append([
            p.get("asin", ""),
            p.get("图片url", ""),
            p.get("标题", ""),
            p.get("详情", ""),
            p.get("核心流量词", ""),
            p.get("俄语标题", ""),
            p.get("俄语详情", ""),
            p.get("货源", ""),       # 空列
            p.get("采购价", ""),     # 空列
            p.get("", ""),           # 空列
            p.get("", ""),           # 空列
            p.get("商品类别", ""),   # 空列
        ])

    wb.save(filepath)
    wb.close()
