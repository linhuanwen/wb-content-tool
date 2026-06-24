"""
AI 图片卡片生成 UI 功能区 — Streamlit Tab4 业务逻辑层。

提供文件校验、生成编排、下载生成等纯数据操作，
供 app.py 的 Streamlit UI 层调用。
"""

import os
import tempfile
from dataclasses import dataclass, field
from pathlib import Path


@dataclass
class CardGenerationUploadResult:
    """卡片生成上传校验结果。

    Attributes:
        is_valid: 校验是否通过。
        error: 失败时的中文错误信息。
        products: 成功时返回的产品列表。
        count: ASIN 数量。
        input_type: "crawler_output"（4列）或 "translation_output"（12列）。
    """
    is_valid: bool = False
    error: str = ""
    products: list[dict] = field(default_factory=list)
    count: int = 0
    input_type: str = ""


def validate_card_generation_upload(filepath: str | Path) -> CardGenerationUploadResult:
    """校验上传的 Excel 文件（用于 AI 图片卡片生成）。

    支持两种格式：
    - 爬虫 4 列输出（asin, 图片url, 标题, 详情）
    - 翻译后 12 列输出（含核心流量词, 俄语标题, 俄语详情 等）

    与 validate_image_upload 逻辑一致，但返回 CardGenerationUploadResult。

    Args:
        filepath: 上传的 Excel 文件路径。

    Returns:
        CardGenerationUploadResult: 校验结果，含产品列表和格式类型。
    """
    filepath = Path(filepath)
    filename = filepath.name

    if not filename.lower().endswith(".xlsx"):
        return CardGenerationUploadResult(
            is_valid=False,
            error="仅支持 .xlsx 格式的 Excel 文件，请重新上传。",
            count=0,
        )

    import openpyxl

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
    except Exception as e:
        return CardGenerationUploadResult(
            is_valid=False,
            error=f"无法打开 Excel 文件: {e}",
            count=0,
        )

    # 读取表头
    headers = [cell.value for cell in ws[1]]
    if headers is None:
        wb.close()
        return CardGenerationUploadResult(
            is_valid=False,
            error="Excel 文件为空或没有表头行。",
            count=0,
        )

    header_set = {str(h).strip() if h is not None else "" for h in headers}

    # 检查必要列
    if "图片url" not in header_set:
        wb.close()
        return CardGenerationUploadResult(
            is_valid=False,
            error="缺少必要列「图片url」。请确保 Excel 包含图片url 列。",
            count=0,
        )

    if "asin" not in header_set:
        wb.close()
        return CardGenerationUploadResult(
            is_valid=False,
            error="缺少必要列「asin」。请确保 Excel 包含 asin 列。",
            count=0,
        )

    # 确定输入类型
    if "俄语标题" in header_set and "核心流量词" in header_set:
        input_type = "translation_output"
    else:
        input_type = "crawler_output"

    # 构建列索引映射
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = str(h).strip() if h is not None else ""
        col_map[key] = i

    # 读取数据行
    products: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        asin = _safe_cell(row, col_map.get("asin"))
        image_urls = _safe_cell(row, col_map.get("图片url"))

        if not asin or not image_urls:
            continue

        product = {"asin": asin, "图片url": image_urls}

        # 构建产品上下文信息库
        product_context: dict[str, str] = {}

        # 始终读取英文原始列（如有）
        for col_name in ("标题", "详情"):
            if col_name in col_map:
                val = _safe_cell(row, col_map[col_name])
                if val:
                    product[col_name] = val
                    product_context[col_name] = val

        # 如果是翻译后 12 列格式，读取俄语列（卡片生成的核心输入）
        if input_type == "translation_output":
            for col_name in ("俄语标题", "核心流量词", "俄语详情"):
                if col_name in col_map:
                    val = _safe_cell(row, col_map[col_name])
                    if val:
                        product_context[col_name] = val

        product["product_context"] = product_context
        products.append(product)

    wb.close()

    return CardGenerationUploadResult(
        is_valid=True,
        error="",
        products=products,
        count=len(products),
        input_type=input_type,
    )


def _safe_cell(row: tuple, col_idx: int | None) -> str:
    """安全读取单元格值。"""
    if col_idx is None:
        return ""
    if col_idx >= len(row):
        return ""
    val = row[col_idx]
    if val is None:
        return ""
    return str(val).strip()


def run_card_generation(
    products: list[dict],
    *,
    mode: str = "card_design",
    custom_prompt: str = "",
) -> "BatchCardResult":
    """同步运行卡片生成（阻塞直到完成），返回 BatchCardResult。

    Args:
        products: 产品列表（需包含 product_context）。
        mode: "card_design"（卡片设计）或 "translate"（图片翻译）。
        custom_prompt: 用户自定义提示词，会附加到 AI 生成 prompt 末尾。

    Returns:
        BatchCardResult。
    """
    from worker import WorkerManager

    manager = WorkerManager()
    return manager.run_card_generation_sync(products=products, mode=mode, custom_prompt=custom_prompt)


def generate_card_output_excel(results: "BatchCardResult") -> bytes:
    """生成 AI 卡片生成的输出 Excel（图片 URL 已替换为 R2 URL）。

    Args:
        results: BatchCardResult 实例。

    Returns:
        xlsx 文件的二进制数据。
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active

    # 表头
    ws.append(["asin", "图片url", "r2卡片url", "状态", "设计说明"])

    for asin_result in results.results:
        asin = asin_result.asin
        for card in asin_result.cards:
            ws.append([
                asin,
                card.original_url,
                card.r2_url,
                card.status,
                card.design_description,
            ])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb.save(tmp_path)
        wb.close()
        with open(tmp_path, "rb") as f:
            return f.read()
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def generate_card_report(results: "BatchCardResult") -> bytes:
    """生成卡片生成报告（CSV 格式）。

    Args:
        results: BatchCardResult 实例。

    Returns:
        CSV 文件的二进制数据（UTF-8 BOM）。
    """
    import csv
    import io

    output = io.StringIO()
    writer = csv.writer(output)

    # 表头
    writer.writerow(["asin", "图片序号", "原始URL", "R2 URL", "状态", "错误信息", "设计说明"])

    for asin_result in results.results:
        asin = asin_result.asin
        for card in asin_result.cards:
            writer.writerow([
                asin,
                card.index,
                card.original_url,
                card.r2_url,
                card.status,
                card.error,
                card.design_description,
            ])

    # UTF-8 BOM（Excel 兼容）
    return ("﻿" + output.getvalue()).encode("utf-8")


def retry_single_card(
    image_url: str,
    asin: str,
    index: int,
    product_context: dict[str, str] | None = None,
    *,
    mode: str = "card_design",
    custom_prompt: str = "",
) -> "CardResult":
    """同步重试单张卡片生成。

    Args:
        image_url: 原始产品图片 URL。
        asin: 产品 ASIN。
        index: 图片序号（0-based）。
        product_context: 产品数据信息库（可选）。
        mode: "card_design"（卡片设计）或 "translate"（图片翻译）。
        custom_prompt: 用户自定义提示词。

    Returns:
        CardResult: 新的生成结果。
    """
    import asyncio

    from image_generator import generate_product_card

    return asyncio.run(
        generate_product_card(
            image_url=image_url,
            asin=asin,
            index=index,
            product_context=product_context,
            mode=mode,
            custom_prompt=custom_prompt,
        )
    )


def _find_product_context(
    products: list[dict], asin: str
) -> dict[str, str] | None:
    """从产品列表中按 ASIN 查找 product_context。"""
    for product in products:
        if product.get("asin") == asin:
            return product.get("product_context", None)
    return None


def _replace_and_recount_cards(
    batch_result: "BatchCardResult",
    asin: str,
    card_index: int,
    new_result: "CardResult",
) -> None:
    """在批量结果中替换单张卡片结果并重算所有计数字段。原地修改 batch_result。

    Args:
        batch_result: 当前的批量处理结果。
        asin: 目标 ASIN。
        card_index: 卡片序号（CardResult.index）。
        new_result: 新的生成结果（retry 产出的 CardResult）。
    """
    # 1. 找到 AsinCardResult，替换 CardResult，累加 retry_count
    for asin_result in batch_result.results:
        if asin_result.asin == asin:
            for i, card in enumerate(asin_result.cards):
                if card.index == card_index:
                    new_result.retry_count = card.retry_count + 1
                    asin_result.cards[i] = new_result
                    break

            # 2. 重算 AsinCardResult 计数
            asin_result.success_count = sum(
                1 for r in asin_result.cards if r.status == "ok"
            )
            asin_result.error_count = sum(
                1 for r in asin_result.cards if r.status == "error"
            )
            asin_result.skipped_count = sum(
                1 for r in asin_result.cards if r.status == "skipped"
            )
            asin_result.video_count = sum(
                1 for r in asin_result.cards if r.status == "video"
            )
            break

    # 3. 重算 BatchCardResult 顶层计数
    batch_result.success_cards = sum(
        r.success_count for r in batch_result.results
    )
    batch_result.error_cards = sum(
        r.error_count for r in batch_result.results
    )
    batch_result.skipped_cards = sum(
        r.skipped_count for r in batch_result.results
    )
    batch_result.video_cards = sum(
        r.video_count for r in batch_result.results
    )
    batch_result.total_cards = sum(
        len(r.cards) for r in batch_result.results
    )


def run_card_generation_from_images(
    uploaded_files: list,
    *,
    custom_prompt: str = "",
    progress_callback=None,
) -> "BatchCardResult":
    """从直接上传的图片文件生成产品卡片。

    每张上传的图片作为独立的产品处理，使用时间戳生成唯一 ASIN。

    Args:
        uploaded_files: Streamlit UploadedFile 对象列表。
        custom_prompt: 用户自定义提示词。
        progress_callback: 每张图片完成后的回调（可选）。

    Returns:
        BatchCardResult。
    """
    import asyncio
    import datetime
    import os
    import tempfile
    import uuid

    from image_generator import AsinCardResult, BatchCardResult, CardResult, generate_card_from_local_image

    started_at = datetime.datetime.now().isoformat()

    # 生成一个批次标识符
    batch_id = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    results: list[AsinCardResult] = []
    total_cards = 0
    success_cards = 0
    error_cards = 0

    async def _process_all():
        nonlocal total_cards, success_cards, error_cards

        for idx, uploaded_file in enumerate(uploaded_files):
            # 保存上传文件到临时路径
            suffix = os.path.splitext(uploaded_file.name)[1] or ".jpg"
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp.write(uploaded_file.getvalue())
                tmp_path = tmp.name

            try:
                # 用文件名（不含扩展名）作为标识
                base_name = os.path.splitext(uploaded_file.name)[0]
                asin = f"UPLOAD_{batch_id}_{base_name[:30]}"
                # 清理 ASIN 中的特殊字符
                import re
                asin = re.sub(r'[^A-Za-z0-9_-]', '_', asin)

                card_result = await generate_card_from_local_image(
                    local_path=tmp_path,
                    asin=asin,
                    index=0,
                    product_context={},
                    mode="card_design",
                    custom_prompt=custom_prompt,
                )
            finally:
                # 清理临时文件
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

            asin_result = AsinCardResult(
                asin=asin,
                cards=[card_result],
                success_count=1 if card_result.status == "ok" else 0,
                error_count=1 if card_result.status == "error" else 0,
            )
            results.append(asin_result)
            total_cards += 1
            if card_result.status == "ok":
                success_cards += 1
            else:
                error_cards += 1

            if progress_callback:
                if asyncio.iscoroutinefunction(progress_callback):
                    await progress_callback(asin_result)
                else:
                    progress_callback(asin_result)

    asyncio.run(_process_all())

    finished_at = datetime.datetime.now().isoformat()

    return BatchCardResult(
        results=results,
        total_asins=len(uploaded_files),
        completed_asins=len(results),
        total_cards=total_cards,
        success_cards=success_cards,
        error_cards=error_cards,
        skipped_cards=0,
        video_cards=0,
        started_at=started_at,
        finished_at=finished_at,
    )


def retry_single_card_from_local(
    local_path: str,
    filename: str,
    *,
    custom_prompt: str = "",
) -> "CardResult":
    """重试单张本地图片的卡片生成。

    Args:
        local_path: 本地图片路径。
        filename: 原始文件名（用于生成 ASIN）。
        custom_prompt: 用户自定义提示词。

    Returns:
        CardResult: 新的生成结果。
    """
    import asyncio
    import datetime

    from image_generator import generate_card_from_local_image

    batch_id = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = os.path.splitext(filename)[0]
    asin = f"RETRY_{batch_id}_{base_name[:30]}"
    import re
    asin = re.sub(r'[^A-Za-z0-9_-]', '_', asin)

    return asyncio.run(
        generate_card_from_local_image(
            local_path=local_path,
            asin=asin,
            index=0,
            product_context={},
            mode="card_design",
            custom_prompt=custom_prompt,
        )
    )
