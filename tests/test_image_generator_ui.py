"""
测试 image_generator_ui.py — AI 图片卡片生成 UI 业务逻辑层。

原则：测试通过公共接口验证行为，mock 外部依赖。
"""

import os
import sys
from unittest.mock import MagicMock, patch

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ============================================================
# 切片1: validate_card_generation_upload — Excel 文件校验
# ============================================================


class TestValidateCardGenerationUpload:
    """上传 Excel 文件校验。"""

    def test_validate_4col_excel_returns_crawler_output(self, tmp_path):
        """上传爬虫 4 列 Excel，返回 input_type="crawler_output"，产品列表正确"""
        import openpyxl

        from image_generator_ui import validate_card_generation_upload

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["asin", "图片url", "标题", "详情"])
        ws.append(["B0TEST1", "https://img1.jpg | https://img2.jpg", "Test Product", "Details"])
        ws.append(["B0TEST2", "https://img3.jpg", "Another Product", "More"])
        filepath = str(tmp_path / "test_4col.xlsx")
        wb.save(filepath)
        wb.close()

        result = validate_card_generation_upload(filepath)

        assert result.is_valid is True
        assert result.input_type == "crawler_output"
        assert result.count == 2
        assert len(result.products) == 2
        assert result.products[0]["asin"] == "B0TEST1"
        assert result.products[0]["图片url"] == "https://img1.jpg | https://img2.jpg"

    def test_validate_12col_excel_returns_translation_output(self, tmp_path):
        """上传翻译后 12 列 Excel，返回 input_type="translation_output"，
        product_context 含俄语字段"""
        import openpyxl

        from image_generator_ui import validate_card_generation_upload

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            "asin", "图片url", "标题", "详情",
            "核心流量词", "俄语标题", "俄语详情",
            "货源", "采购价", "", "", "商品类别",
        ])
        ws.append([
            "B0T12COL", "https://img1.jpg", "Product X", "Details X",
            "ключевые слова", "Название продукта", "Описание товара",
            "", "", "", "", "",
        ])
        filepath = str(tmp_path / "test_12col.xlsx")
        wb.save(filepath)
        wb.close()

        result = validate_card_generation_upload(filepath)

        assert result.is_valid is True
        assert result.input_type == "translation_output"
        assert result.count == 1
        ctx = result.products[0]["product_context"]
        assert ctx["俄语标题"] == "Название продукта"
        assert ctx["核心流量词"] == "ключевые слова"
        assert ctx["俄语详情"] == "Описание товара"

    def test_validate_rejects_missing_asin_column(self, tmp_path):
        """缺少 asin 列 → is_valid=False"""
        import openpyxl

        from image_generator_ui import validate_card_generation_upload

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["图片url", "标题", "详情"])
        ws.append(["https://img1.jpg", "Product", "Details"])
        filepath = str(tmp_path / "test_missing_asin.xlsx")
        wb.save(filepath)
        wb.close()

        result = validate_card_generation_upload(filepath)

        assert result.is_valid is False
        assert "asin" in result.error.lower()

    def test_validate_rejects_non_xlsx(self, tmp_path):
        """.csv 文件 → is_valid=False"""
        from image_generator_ui import validate_card_generation_upload

        filepath = str(tmp_path / "test.csv")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("asin,图片url,标题,详情\nB01,url,Title,Details\n")

        result = validate_card_generation_upload(filepath)

        assert result.is_valid is False
        assert "xlsx" in result.error.lower()


# ============================================================
# 切片2: generate_card_output_excel + generate_card_report — 下载生成
# ============================================================


class TestGenerateCardOutput:
    """输出 Excel 和报告生成。"""

    def test_generate_output_excel_contains_r2_urls(self):
        """验证输出 xlsx 包含 R2 卡片 URL 和设计说明列"""
        from image_generator import AsinCardResult, BatchCardResult, CardResult
        from image_generator_ui import generate_card_output_excel

        cards = [
            CardResult(
                index=0,
                original_url="https://amazon.com/img1.jpg",
                r2_url="https://pub-xxx.r2.dev/B01/00_card.jpg",
                status="ok",
                design_description="现代厨房场景，暖色调",
            ),
            CardResult(
                index=1,
                original_url="https://amazon.com/img2.jpg",
                r2_url="",
                status="error",
                error="Gemini 生成失败",
            ),
        ]
        results = BatchCardResult(
            results=[
                AsinCardResult(asin="B01", cards=cards, success_count=1, error_count=1),
            ],
            total_asins=1,
            completed_asins=1,
            total_cards=2,
            success_cards=1,
            error_cards=1,
        )

        xlsx_bytes = generate_card_output_excel(results)

        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 0

        import io
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        assert "r2卡片url" in headers
        assert "设计说明" in headers

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 2

    def test_generate_report_contains_all_statuses(self):
        """报告含 ok/error/skipped 三类卡片，验证状态和 ASIN"""
        from image_generator import AsinCardResult, BatchCardResult, CardResult
        from image_generator_ui import generate_card_report

        cards = [
            CardResult(index=0, original_url="url1", r2_url="r2_1", status="ok", design_description="设计1"),
            CardResult(index=1, original_url="url2", r2_url="r2_2", status="error", error="生成失败"),
            CardResult(index=2, original_url="url3", r2_url="", status="skipped"),
        ]
        results = BatchCardResult(
            results=[AsinCardResult(asin="B0REPORT", cards=cards, success_count=1, error_count=1, skipped_count=1)],
            total_asins=1, completed_asins=1, total_cards=3,
            success_cards=1, error_cards=1, skipped_cards=1,
        )

        csv_bytes = generate_card_report(results)

        assert isinstance(csv_bytes, bytes)
        report_text = csv_bytes.decode("utf-8")

        assert "ok" in report_text
        assert "error" in report_text
        assert "skipped" in report_text
        assert "B0REPORT" in report_text
        assert "设计1" in report_text


# ============================================================
# 切片3: retry_single_card — 单卡片重试
# ============================================================


class TestRetrySingleCard:
    """retry_single_card 同步包装器。"""

    def test_retry_single_card_returns_card_result(self):
        """retry_single_card 返回有效的 CardResult"""
        from image_generator_ui import retry_single_card

        fake_bytes = _fake_image_bytes()

        async def mock_download(url):
            return "/tmp/test.jpg"

        async def mock_generate(image_path, product_context=None, *, mode="card_design", **kwargs):
            return fake_bytes, "简约设计，白色背景"

        async def mock_resize(image):
            return image

        async def mock_upload(local_path, remote_key):
            return f"https://pub-xxx.r2.dev/{remote_key}"

        from unittest.mock import patch

        with (
            patch("image_generator._real_download", side_effect=mock_download),
            patch("image_generator._dispatch_generate", side_effect=mock_generate),
            patch("image_generator._real_upload", side_effect=mock_upload),
        ):
            result = retry_single_card(
                image_url="https://example.com/product.jpg",
                asin="B0TEST",
                index=0,
                product_context={"俄语标题": "тест"},
            )

        assert result.status == "ok"
        assert result.original_url == "https://example.com/product.jpg"
        assert "简约设计" in result.design_description


def _fake_image_bytes() -> bytes:
    from io import BytesIO
    from PIL import Image

    buf = BytesIO()
    img = Image.new("RGB", (900, 1200), color=(200, 180, 160))
    img.save(buf, "JPEG", quality=95)
    return buf.getvalue()


# ============================================================
# 切片4: _replace_and_recount_cards — 原地替换+计数重算
# ============================================================


class TestReplaceAndRecountCards:
    """_replace_and_recount_cards 原地替换 CardResult 并重算所有计数。"""

    def test_replace_error_to_ok_updates_counts(self):
        """error→ok 替换后，Asin 和 Batch 计数都正确更新"""
        from image_generator import AsinCardResult, BatchCardResult, CardResult
        from image_generator_ui import _replace_and_recount_cards

        old = CardResult(index=0, original_url="url1", status="error", error="生成失败")
        batch = BatchCardResult(
            results=[AsinCardResult(asin="B0X", cards=[old], success_count=0, error_count=1, skipped_count=0)],
            total_asins=1, completed_asins=1, total_cards=1,
            success_cards=0, error_cards=1, skipped_cards=0,
        )

        new = CardResult(index=0, original_url="url1", r2_url="https://r2.dev/B0X/00_card.jpg", status="ok")

        _replace_and_recount_cards(batch, "B0X", 0, new)

        # Asin 级
        assert batch.results[0].success_count == 1
        assert batch.results[0].error_count == 0
        # Batch 级
        assert batch.success_cards == 1
        assert batch.error_cards == 0
        assert batch.total_cards == 1

    def test_replace_increments_retry_count(self):
        """替换后 retry_count 在原值基础上 +1"""
        from image_generator import AsinCardResult, BatchCardResult, CardResult
        from image_generator_ui import _replace_and_recount_cards

        old = CardResult(index=0, original_url="url1", status="error", retry_count=2)
        batch = BatchCardResult(
            results=[AsinCardResult(asin="B0X", cards=[old], success_count=0, error_count=1, skipped_count=0)],
            total_cards=1, success_cards=0, error_cards=1, skipped_cards=0,
        )

        new = CardResult(index=0, original_url="url1", status="ok")
        _replace_and_recount_cards(batch, "B0X", 0, new)

        assert batch.results[0].cards[0].retry_count == 3  # 2 + 1
