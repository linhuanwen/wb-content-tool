"""
测试 image_translator_ui.py — 图片翻译 UI 业务逻辑层。

原则：测试通过公共接口验证行为，mock 外部依赖。
"""

import os
import sys
from unittest.mock import MagicMock, patch

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


@pytest.fixture(autouse=True)
def _reset_worker_manager():
    """每个测试前后重置 WorkerManager 单例，避免测试间状态泄漏。"""
    from worker import WorkerManager
    WorkerManager.reset()
    yield
    WorkerManager.reset()


# ============================================================
# 切片0: enrich_product_context_from_db — 从 DB 丰富产品上下文
# ============================================================


class TestEnrichProductContextFromDb:
    """enrich_product_context_from_db 从数据库读取 Phase 1/2 产物。"""

    def test_enriches_with_phase1_product_info(self, tmp_path):
        """Phase 1 完成后，产品上下文含英文属性"""
        import sqlite3

        from image_translator_ui import enrich_product_context_from_db
        from db import init_db, upsert_product

        db_path = str(tmp_path / "products.db")
        db = sqlite3.connect(db_path)
        init_db(db)
        upsert_product(db, {
            "asin": "B0TESTDB",
            "title": "Test Product Title",
            "details": "Test product details",
            "image_urls": "https://img1.jpg;https://img2.jpg",
            "category": "Electronics",
            "material": "Plastic",
            "features": ["Feature A", "Feature B", "Feature C"],
            "target_audience": "Adults",
            "brand": "TestBrand",
        })
        db.close()

        products = [{"asin": "B0TESTDB", "图片url": "", "product_context": {}}]
        enriched = enrich_product_context_from_db(products, db_path)

        ctx = enriched[0]["product_context"]
        assert ctx["标题"] == "Test Product Title"
        assert ctx["详情"] == "Test product details"
        assert ctx["product_category"] == "Electronics"
        assert ctx["material"] == "Plastic"
        assert "Feature A" in ctx["features"]
        assert ctx["target_audience"] == "Adults"
        assert ctx["brand"] == "TestBrand"

    def test_enriches_with_phase2_translation(self, tmp_path):
        """Phase 2 完成后，产品上下文含俄语文案"""
        import sqlite3

        from image_translator_ui import enrich_product_context_from_db
        from db import init_db, upsert_product, upsert_translation

        db_path = str(tmp_path / "products.db")
        db = sqlite3.connect(db_path)
        init_db(db)
        upsert_product(db, {
            "asin": "B0TRANS",
            "title": "Product",
            "details": "Details",
            "image_urls": "https://img1.jpg",
        })
        upsert_translation(db, {
            "asin": "B0TRANS",
            "russian_title": "Название продукта",
            "core_keywords": "ключевые слова",
            "russian_description": "Описание продукта",
        })
        db.close()

        products = [{"asin": "B0TRANS", "图片url": "", "product_context": {}}]
        enriched = enrich_product_context_from_db(products, db_path)

        ctx = enriched[0]["product_context"]
        assert ctx["俄语标题"] == "Название продукта"
        assert ctx["核心流量词"] == "ключевые слова"
        assert ctx["俄语详情"] == "Описание продукта"

    def test_db_values_override_excel_values(self, tmp_path):
        """数据库中的字段值覆盖 Excel 中的同名字段"""
        import sqlite3

        from image_translator_ui import enrich_product_context_from_db
        from db import init_db, upsert_product, upsert_translation

        db_path = str(tmp_path / "products.db")
        db = sqlite3.connect(db_path)
        init_db(db)
        upsert_product(db, {
            "asin": "B0OVERRIDE",
            "title": "DB Title",
            "details": "DB Details",
        })
        upsert_translation(db, {
            "asin": "B0OVERRIDE",
            "russian_title": "DB Русский",
            "core_keywords": "DB Keywords",
            "russian_description": "DB Описание",
        })
        db.close()

        # Excel 中有旧值，DB 中有新值
        products = [{
            "asin": "B0OVERRIDE",
            "图片url": "",
            "product_context": {
                "标题": "Old Excel Title",
                "俄语标题": "Old Excel Russian",
            },
        }]
        enriched = enrich_product_context_from_db(products, db_path)

        ctx = enriched[0]["product_context"]
        # DB 值覆盖 Excel 值
        assert ctx["标题"] == "DB Title"
        assert ctx["俄语标题"] == "DB Русский"

    def test_missing_db_does_not_crash(self):
        """数据库文件不存在时不崩溃，返回原 products"""
        from image_translator_ui import enrich_product_context_from_db

        products = [{"asin": "B0NODB", "product_context": {"标题": "Original"}}]
        enriched = enrich_product_context_from_db(products, "nonexistent.db")

        # 原产品不被修改
        assert enriched[0]["product_context"]["标题"] == "Original"

    def test_returns_same_list_for_chaining(self):
        """返回同一列表，支持链式调用"""
        from image_translator_ui import enrich_product_context_from_db

        products = [{"asin": "B0CHAIN", "product_context": {}}]
        result = enrich_product_context_from_db(products, "nonexistent.db")

        assert result is products

    def test_phase2_not_complete_only_provides_phase1(self, tmp_path):
        """Phase 2 未完成时仅提供 Phase 1 英文属性"""
        import sqlite3

        from image_translator_ui import enrich_product_context_from_db
        from db import init_db, upsert_product

        db_path = str(tmp_path / "products.db")
        db = sqlite3.connect(db_path)
        init_db(db)
        upsert_product(db, {
            "asin": "B0P1ONLY",
            "title": "Phase1 Title",
            "details": "Phase1 Details",
        })
        db.close()

        products = [{"asin": "B0P1ONLY", "product_context": {}}]
        enriched = enrich_product_context_from_db(products, db_path)

        ctx = enriched[0]["product_context"]
        # Phase 1 字段存在
        assert ctx["标题"] == "Phase1 Title"
        # Phase 2 字段不存在（翻译未完成）
        assert "俄语标题" not in ctx or ctx.get("俄语标题") == ""


# ============================================================
# 切片1: validate_image_upload — Excel 文件校验
# ============================================================


class TestValidateImageUpload:
    """上传 Excel 文件校验。"""

    def test_validate_4col_excel_returns_crawler_output(self, tmp_path):
        """上传爬虫 4 列 Excel，返回 input_type="crawler_output"，产品列表正确"""
        import openpyxl

        from image_translator_ui import validate_image_upload

        # 创建 4 列 Excel（爬虫输出格式）
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["asin", "图片url", "标题", "详情"])
        ws.append(["B0TEST1", "https://img1.jpg | https://img2.jpg", "Test Product", "Details here"])
        ws.append(["B0TEST2", "https://img3.jpg", "Another Product", "More details"])
        filepath = str(tmp_path / "test_4col.xlsx")
        wb.save(filepath)
        wb.close()

        result = validate_image_upload(filepath)

        assert result.is_valid is True
        assert result.input_type == "crawler_output"
        assert result.count == 2
        assert len(result.products) == 2
        assert result.products[0]["asin"] == "B0TEST1"
        assert result.products[0]["图片url"] == "https://img1.jpg | https://img2.jpg"

    def test_validate_12col_excel_returns_translation_output(self, tmp_path):
        """上传翻译后 12 列 Excel，返回 input_type="translation_output"，产品列表正确"""
        import openpyxl

        from image_translator_ui import validate_image_upload

        # 创建 12 列 Excel（翻译后格式）
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            "asin", "图片url", "标题", "详情",
            "核心流量词", "俄语标题", "俄语详情",
            "货源", "采购价", "", "", "商品类别",
        ])
        ws.append([
            "B0T12COL", "https://img1.jpg", "Product X", "Details X",
            "ключевые слова", "Название", "Описание",
            "", "", "", "", "",
        ])
        filepath = str(tmp_path / "test_12col.xlsx")
        wb.save(filepath)
        wb.close()

        result = validate_image_upload(filepath)

        assert result.is_valid is True
        assert result.input_type == "translation_output"
        assert result.count == 1
        assert len(result.products) == 1
        assert result.products[0]["asin"] == "B0T12COL"

    def test_validate_rejects_missing_image_url_column(self, tmp_path):
        """缺少 图片url 列，返回 is_valid=False + 错误信息"""
        import openpyxl

        from image_translator_ui import validate_image_upload

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["asin", "标题", "详情"])  # 缺 图片url 列
        ws.append(["B0TEST1", "Product", "Details"])
        filepath = str(tmp_path / "test_missing_col.xlsx")
        wb.save(filepath)
        wb.close()

        result = validate_image_upload(filepath)

        assert result.is_valid is False
        assert result.error != ""
        assert "图片url" in result.error
        assert result.count == 0

    def test_validate_rejects_non_xlsx(self, tmp_path):
        """.csv 文件，返回 is_valid=False"""
        from image_translator_ui import validate_image_upload

        filepath = str(tmp_path / "test.csv")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("asin,图片url,标题,详情\nB01,url,Title,Details\n")

        result = validate_image_upload(filepath)

        assert result.is_valid is False
        assert "xlsx" in result.error.lower()
        assert result.count == 0


# ============================================================
# 切片2: generate_output_excel + generate_report — 下载生成
# ============================================================


class TestGenerateOutput:
    """输出 Excel 和报告生成。"""

    def test_generate_output_excel_urls_replaced(self):
        """验证输出 xlsx 中 r2图片url 列包含 R2 URL"""
        from image_translator import AsinImageResult, BatchImageResult, ImageResult
        from image_translator_ui import generate_output_excel

        # 构造 BatchImageResult
        images = [
            ImageResult(
                index=0,
                original_url="https://amazon.com/img1.jpg",
                r2_url="https://pub-xxx.r2.dev/B01/00_ru.jpg",
                status="ok",
            ),
            ImageResult(
                index=1,
                original_url="https://amazon.com/img2.jpg",
                r2_url="https://pub-xxx.r2.dev/B01/01_ru.jpg",
                status="skipped",
                error="OCR 失败",
            ),
        ]
        results = BatchImageResult(
            results=[
                AsinImageResult(
                    asin="B01",
                    images=images,
                    success_count=1,
                    error_count=0,
                    skipped_count=1,
                ),
            ],
            total_asins=1,
            completed_asins=1,
            total_images=2,
            success_images=1,
            skipped_images=1,
        )

        xlsx_bytes = generate_output_excel(results)

        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 0

        # 验证内容：用 openpyxl 重新读取
        import io
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active

        # 表头
        headers = [cell.value for cell in ws[1]]
        assert "r2图片url" in headers

        # 数据行
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 2  # 2 张图片

        # 验证 R2 URL 已写入
        r2_col = headers.index("r2图片url")
        row0_r2 = rows[0][r2_col]
        assert "r2.dev" in row0_r2

    def test_generate_report_contains_all_statuses(self):
        """报告含 ok/skipped/error 三类图片，验证每行列出了状态和 asin"""
        from image_translator import AsinImageResult, BatchImageResult, ImageResult
        from image_translator_ui import generate_report

        images = [
            ImageResult(index=0, original_url="url1", r2_url="r2_1", status="ok"),
            ImageResult(index=1, original_url="url2", r2_url="r2_2", status="skipped", error="OCR 失败"),
            ImageResult(index=2, original_url="url3", r2_url="url3", status="error", error="下载失败"),
        ]
        results = BatchImageResult(
            results=[
                AsinImageResult(asin="B0REPORT", images=images, success_count=1, error_count=1, skipped_count=1),
            ],
            total_asins=1,
            completed_asins=1,
            total_images=3,
            success_images=1,
            error_images=1,
            skipped_images=1,
        )

        csv_bytes = generate_report(results)

        assert isinstance(csv_bytes, bytes)
        report_text = csv_bytes.decode("utf-8")

        # 验证含所有三种状态
        assert "ok" in report_text
        assert "skipped" in report_text
        assert "error" in report_text
        assert "B0REPORT" in report_text
        # 验证错误信息列
        assert "OCR 失败" in report_text
        assert "下载失败" in report_text


# ============================================================
# 切片3: Worker 集成 — start/get_status/pause/resume
# ============================================================


class TestWorkerIntegration:
    """start_background_translation / get_worker_status / pause_resume。"""

    def test_get_worker_status_returns_idle_initially(self):
        """初始状态下 get_worker_status() 返回 idle"""
        from image_translator_ui import get_worker_status

        status = get_worker_status()
        assert status.state == "idle"

    def test_pause_and_resume_worker(self):
        """暂停后恢复 Worker"""
        from image_translator_ui import pause_worker, resume_worker

        # 暂停和恢复不应抛异常
        pause_worker()
        resume_worker()

    def test_start_background_translation_starts_worker(self):
        """start_background_translation 启动 Worker 并变更状态"""
        from image_processor import FontConfig, TextRegion
        from image_translator_ui import get_worker_status, start_background_translation

        # mock 函数让 Worker 快速完成
        async def mock_download(url):
            return "/tmp/test.jpg"

        async def mock_ocr(local_path):
            return [TextRegion(text="Test", translation="", box=(10, 20, 100, 50))]

        async def mock_translate(texts):
            return ["тест"]

        async def mock_repair(image, regions):
            return image

        async def mock_resize(image):
            from PIL import Image
            return Image.new("RGB", (900, 1200))

        async def mock_upload(local_path, remote_key):
            return f"https://pub-xxx.r2.dev/{remote_key}"

        products = [
            {"asin": "B0UIINT", "图片url": "https://img1.jpg"},
        ]

        start_background_translation(
            products=products,
            font_config=FontConfig(),
        )

        # Worker 可能已经完成（很快），也可能还在运行
        import time
        time.sleep(0.3)

        status = get_worker_status()
        # 状态应为 running 或 completed
        assert status.state in ("running", "completed"), f"实际状态: {status.state}"


# ============================================================
# 切片4: retry_single_image — 单图重试
# ============================================================


class TestRetrySingleImage:
    """retry_single_image 同步包装器。"""

    def test_retry_single_image_returns_image_result(self):
        """retry_single_image 返回有效的 ImageResult"""
        from image_processor import FontConfig, TextRegion
        from image_translator_ui import retry_single_image

        # 构建轻量 mock 管线，绕过真实下载/OCR/上传
        async def mock_download(url):
            return "/tmp/test.jpg"

        async def mock_ocr(local_path):
            return [TextRegion(text="Hello", translation="", box=(10, 20, 100, 50))]

        async def mock_translate(texts, product_context=None):
            return ["привет"]

        async def mock_repair(image, regions):
            return image

        async def mock_resize(image):
            from PIL import Image
            return Image.new("RGB", (900, 1200))

        async def mock_upload(local_path, remote_key):
            return f"https://pub-xxx.r2.dev/{remote_key}"

        from unittest.mock import patch

        with patch("image_translator._real_download", side_effect=mock_download), \
             patch("image_translator._real_ocr", side_effect=mock_ocr), \
             patch("image_translator._real_translate", side_effect=mock_translate), \
             patch("image_translator._real_upload", side_effect=mock_upload):
            result = retry_single_image(
                image_url="https://example.com/img.jpg",
                asin="B0TEST",
                index=0,
                font_config=FontConfig(),
            )

        assert result.status == "ok"
        assert result.original_url == "https://example.com/img.jpg"
        assert len(result.translated_texts) == 1
        assert result.translated_texts[0] == "привет"

    def test_retry_single_image_passes_product_context(self):
        """retry_single_image 将 product_context 传递给翻译函数"""
        from image_processor import FontConfig, TextRegion
        from image_translator_ui import retry_single_image

        received_context = []

        async def mock_download(url):
            return "/tmp/test.jpg"

        async def mock_ocr(local_path):
            return [TextRegion(text="Hello", translation="", box=(10, 20, 100, 50))]

        async def mock_translate(texts, product_context=None):
            received_context.append(product_context)
            return ["привет"]

        async def mock_repair(image, regions):
            return image

        async def mock_resize(image):
            from PIL import Image
            return Image.new("RGB", (900, 1200))

        async def mock_upload(local_path, remote_key):
            return f"https://pub-xxx.r2.dev/{remote_key}"

        from unittest.mock import patch

        ctx = {"俄语标题": "тест", "核心流量词": "ключ"}

        with patch("image_translator._real_download", side_effect=mock_download), \
             patch("image_translator._real_ocr", side_effect=mock_ocr), \
             patch("image_translator._real_translate", side_effect=mock_translate), \
             patch("image_translator._real_upload", side_effect=mock_upload):
            retry_single_image(
                image_url="https://example.com/img.jpg",
                asin="B0TEST",
                index=0,
                font_config=FontConfig(),
                product_context=ctx,
            )

        assert len(received_context) == 1
        assert received_context[0] == ctx


# ============================================================
# 切片5: _replace_and_recount — 原地替换+计数重算
# ============================================================


class TestReplaceAndRecount:
    """_replace_and_recount 原地替换 ImageResult 并重算所有计数。"""

    def test_replace_error_to_ok_updates_counts(self):
        """error→ok 替换后，Asin 和 Batch 计数都正确更新"""
        from image_translator import AsinImageResult, BatchImageResult, ImageResult
        from image_translator_ui import _replace_and_recount

        old = ImageResult(
            index=0,
            original_url="url1",
            status="error",
            error="翻译失败",
        )
        batch = BatchImageResult(
            results=[
                AsinImageResult(
                    asin="B0X",
                    images=[old],
                    success_count=0,
                    error_count=1,
                    skipped_count=0,
                ),
            ],
            total_asins=1,
            completed_asins=1,
            total_images=1,
            success_images=0,
            error_images=1,
            skipped_images=0,
        )

        new = ImageResult(
            index=0,
            original_url="url1",
            r2_url="https://r2.dev/B0X/00_ru.jpg",
            status="ok",
        )

        _replace_and_recount(batch, "B0X", 0, new)

        # Asin 级
        assert batch.results[0].success_count == 1
        assert batch.results[0].error_count == 0
        assert batch.results[0].skipped_count == 0
        # Batch 级
        assert batch.success_images == 1
        assert batch.error_images == 0
        assert batch.skipped_images == 0
        assert batch.total_images == 1

    def test_replace_skipped_to_ok_updates_counts(self):
        """skipped→ok 替换后计数正确"""
        from image_translator import AsinImageResult, BatchImageResult, ImageResult
        from image_translator_ui import _replace_and_recount

        old = ImageResult(
            index=0,
            original_url="url1",
            status="skipped",
            error="OCR 失败",
        )
        batch = BatchImageResult(
            results=[
                AsinImageResult(
                    asin="B0X",
                    images=[old],
                    success_count=0,
                    error_count=0,
                    skipped_count=1,
                ),
            ],
            total_images=1,
            success_images=0,
            error_images=0,
            skipped_images=1,
        )

        new = ImageResult(index=0, original_url="url1", status="ok")

        _replace_and_recount(batch, "B0X", 0, new)

        assert batch.results[0].success_count == 1
        assert batch.results[0].skipped_count == 0
        assert batch.success_images == 1
        assert batch.skipped_images == 0

    def test_replace_increments_retry_count(self):
        """替换后 retry_count 在原值基础上 +1"""
        from image_translator import AsinImageResult, BatchImageResult, ImageResult
        from image_translator_ui import _replace_and_recount

        old = ImageResult(
            index=0,
            original_url="url1",
            status="error",
            retry_count=2,
        )
        batch = BatchImageResult(
            results=[
                AsinImageResult(
                    asin="B0X",
                    images=[old],
                    success_count=0,
                    error_count=1,
                    skipped_count=0,
                ),
            ],
            total_images=1,
            success_images=0,
            error_images=1,
            skipped_images=0,
        )

        new = ImageResult(index=0, original_url="url1", status="ok")

        _replace_and_recount(batch, "B0X", 0, new)

        assert batch.results[0].images[0].retry_count == 3  # 2 + 1

    def test_replace_multiple_asins_only_affects_target(self):
        """两个 ASIN（各有 1 error），只重试一个，另一个不变"""
        from image_translator import AsinImageResult, BatchImageResult, ImageResult
        from image_translator_ui import _replace_and_recount

        batch = BatchImageResult(
            results=[
                AsinImageResult(
                    asin="B0A",
                    images=[
                        ImageResult(index=0, original_url="urlA", status="error", error="E1"),
                    ],
                    success_count=0,
                    error_count=1,
                    skipped_count=0,
                ),
                AsinImageResult(
                    asin="B0B",
                    images=[
                        ImageResult(index=0, original_url="urlB", status="error", error="E2"),
                    ],
                    success_count=0,
                    error_count=1,
                    skipped_count=0,
                ),
            ],
            total_images=2,
            success_images=0,
            error_images=2,
            skipped_images=0,
        )

        new = ImageResult(index=0, original_url="urlA", status="ok")

        _replace_and_recount(batch, "B0A", 0, new)

        # B0A 更新了
        assert batch.results[0].success_count == 1
        assert batch.results[0].error_count == 0
        # B0B 未变
        assert batch.results[1].success_count == 0
        assert batch.results[1].error_count == 1
        # Batch 级：1 ok + 1 error
        assert batch.success_images == 1
        assert batch.error_images == 1
        assert batch.total_images == 2


# ============================================================
# 切片6: _find_product_context — 按 ASIN 查找上下文
# ============================================================


class TestFindProductContext:
    """_find_product_context 从产品列表查找 product_context。"""

    def test_find_existing_asin_returns_context(self):
        """ASIN 存在时返回正确的 product_context"""
        from image_translator_ui import _find_product_context

        ctx = {"标题": "Test", "俄语标题": "тест"}
        products = [
            {"asin": "B01", "product_context": {"标题": "Other"}},
            {"asin": "B02", "product_context": ctx},
            {"asin": "B03", "product_context": {}},
        ]

        result = _find_product_context(products, "B02")
        assert result == ctx

    def test_find_missing_asin_returns_none(self):
        """ASIN 不存在时返回 None"""
        from image_translator_ui import _find_product_context

        products = [
            {"asin": "B01", "product_context": {"标题": "Test"}},
        ]

        result = _find_product_context(products, "B99")
        assert result is None
