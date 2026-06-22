"""
测试 crawler_ui.py — Web 爬虫功能区业务逻辑。

测试行为（非实现）：
- 文件上传校验（扩展名、ASIN 列检测）
- 采集编排（正常流程 + 失败收集）
- 下载文件生成
"""

import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ============================================================
# 测试辅助函数
# ============================================================

def _create_excel(path: str, headers: list[str], rows: list[list]) -> None:
    """创建测试用 Excel 文件。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _make_temp_xlsx(headers: list[str], rows: list[list]) -> str:
    """创建一个临时 Excel 文件并返回路径（调用方负责清理）。"""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    _create_excel(path, headers, rows)
    return path


def _make_temp_csv() -> str:
    """创建一个临时 .csv 文件并返回路径（调用方负责清理）。"""
    fd, path = tempfile.mkstemp(suffix=".csv")
    os.close(fd)
    with open(path, "w", encoding="utf-8") as f:
        f.write("dummy,data\n")
    return path


def _safe_unlink(path: str) -> None:
    """安全删除临时文件。"""
    try:
        os.unlink(path)
    except OSError:
        pass


# ============================================================
# validate_upload 测试
# ============================================================

class TestValidateUpload:
    """上传文件校验行为。"""

    # --- 切片 1：扩展名校验 ---

    def test_rejects_non_xlsx_file(self):
        """上传 .csv 文件时返回友好错误提示"""
        from crawler_ui import validate_upload

        csv_path = _make_temp_csv()
        try:
            result = validate_upload(csv_path)
            assert result.is_valid is False
            assert ".xlsx" in result.error.lower() or "Excel" in result.error
            assert result.asins == []
            assert result.count == 0
        finally:
            _safe_unlink(csv_path)

    # --- 切片 2：ASIN 列缺失校验 ---

    def test_rejects_excel_without_asin_column(self):
        """上传有效 .xlsx 但没有 asin 列时返回错误"""
        from crawler_ui import validate_upload

        path = _make_temp_xlsx(
            headers=["商品编号", "标题"],
            rows=[["12345", "Product A"]],
        )
        try:
            result = validate_upload(path)
            assert result.is_valid is False
            assert "asin" in result.error.lower()
            assert result.asins == []
            assert result.count == 0
        finally:
            _safe_unlink(path)

    # --- 切片 3：成功上传 —— ASIN 识别 + 计数 ---

    def test_returns_asins_and_count_for_valid_excel(self):
        """上传含 asin 列的有效 Excel → 返回 ASIN 列表和正确计数"""
        from crawler_ui import validate_upload

        path = _make_temp_xlsx(
            headers=["asin", "标题", "详情"],
            rows=[
                ["B0GVYXC124", "Product A", "Description A"],
                ["B0F45N6NS7", "Product B", "Description B"],
                ["B000000000", "Product C", "Description C"],
            ],
        )
        try:
            result = validate_upload(path)
            assert result.is_valid is True
            assert result.error == ""
            assert result.asins == ["B0GVYXC124", "B0F45N6NS7", "B000000000"]
            assert result.count == 3
        finally:
            _safe_unlink(path)


# ============================================================
# execute_crawl 测试
# ============================================================

# 用与 test_crawler.py 相同的 HTML fixture
FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "fixtures")


def _load_fixture(filename: str) -> str:
    path = os.path.join(FIXTURES_DIR, filename)
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


VALID_PRODUCT_HTML = _load_fixture("amazon_product_sample.html")


class TestExecuteCrawl:
    """采集编排行为。"""

    @pytest.mark.asyncio
    async def test_crawls_all_asins_and_returns_crawl_result(self):
        """正常流程：所有 ASIN 成功 → 返回 CrawlResult"""
        from crawler_ui import execute_crawl

        async def fake_fetch(asin):
            return VALID_PRODUCT_HTML

        result = await execute_crawl(
            ["B0A", "B0B", "B0C"],
            _fetch_func=fake_fetch,
            delay_min=0,
            delay_max=0,
        )

        assert len(result.products) == 3
        assert result.failed_asins == []
        assert result.products[0]["asin"] in ("B0A", "B0B", "B0C")

    @pytest.mark.asyncio
    async def test_collects_failed_asins(self):
        """部分失败：成功的入 products，失败的入 failed_asins"""
        from crawler_ui import execute_crawl

        async def flaky_fetch(asin):
            if asin == "B0_BAD":
                raise RuntimeError("Network error")
            return VALID_PRODUCT_HTML

        result = await execute_crawl(
            ["B0_GOOD1", "B0_BAD", "B0_GOOD2"],
            _fetch_func=flaky_fetch,
            delay_min=0,
            delay_max=0,
            max_retries=0,  # 不重试，立即失败
        )

        assert len(result.products) == 2
        assert result.failed_asins == ["B0_BAD"]
        for p in result.products:
            assert p["asin"] != "B0_BAD"

    @pytest.mark.asyncio
    async def test_all_failed_returns_empty_products(self):
        """全部失败：products 为空，全部 ASIN 在 failed_asins"""
        from crawler_ui import execute_crawl

        async def always_fail(asin):
            raise RuntimeError("Boom")

        result = await execute_crawl(
            ["B0A", "B0B"],
            _fetch_func=always_fail,
            delay_min=0,
            delay_max=0,
            max_retries=0,
        )

        assert result.products == []
        assert set(result.failed_asins) == {"B0A", "B0B"}

    @pytest.mark.asyncio
    async def test_progress_callback_fires(self):
        """进度回调在每次完成时触发"""
        from crawler_ui import execute_crawl

        progress_entries = []

        def on_progress(current, total, asin, status):
            progress_entries.append((current, total, asin, status))

        async def fake_fetch(asin):
            return VALID_PRODUCT_HTML

        await execute_crawl(
            ["B0A", "B0B"],
            _fetch_func=fake_fetch,
            delay_min=0,
            delay_max=0,
            progress_callback=on_progress,
        )

        assert len(progress_entries) == 2
        assert all(status == "ok" for _, _, _, status in progress_entries)


# ============================================================
# generate_download 测试
# ============================================================

SAMPLE_PRODUCTS = [
    {
        "asin": "B0GVYXC124",
        "图片url": "https://img1.jpg;https://img2.jpg",
        "标题": "5-in-1 Face & Body Sculpting Machine",
        "详情": "Multifunctional body contouring machine.",
    },
    {
        "asin": "B0F45N6NS7",
        "图片url": "https://img3.jpg",
        "标题": "Anti Cellulite Massager",
        "详情": "Handheld electric anti-cellulite device.",
    },
]


class TestGenerateDownload:
    """下载文件生成行为。"""

    def test_returns_valid_xlsx_bytes(self):
        """返回非空的 xlsx 字节流，可被 openpyxl 重新打开"""
        from crawler_ui import generate_download

        data = generate_download(SAMPLE_PRODUCTS)

        assert isinstance(data, bytes)
        assert len(data) > 0

        # 验证可以被 openpyxl 重新读取
        import openpyxl
        import io

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        wb.close()

        assert headers == ["asin", "图片url", "标题", "详情"]

    def test_rows_match_input_products(self):
        """下载的 xlsx 数据行与输入产品一致"""
        from crawler_ui import generate_download

        data = generate_download(SAMPLE_PRODUCTS)

        import openpyxl
        import io

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        assert len(rows) == 2
        assert rows[0][0] == "B0GVYXC124"
        assert rows[1][0] == "B0F45N6NS7"
