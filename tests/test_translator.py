"""
测试 translator.py — AI 翻译模块。

测试行为（非实现）：
- TranslationProvider 是抽象基类，不能直接实例化
- 工厂函数根据配置创建正确的 Provider 实例
- 标题后处理规则
- JSON 解析与重试
- 批量翻译端到端行为
- CLI 入口
"""

import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ============================================================
# 切片1: TranslationProvider ABC + 工厂函数
# ============================================================

class TestTranslationProviderIsAbstract:
    """TranslationProvider 是抽象基类，不能直接实例化。"""

    def test_cannot_instantiate_abstract_provider(self):
        """直接实例化 TranslationProvider 应抛出 TypeError"""
        from translator import TranslationProvider

        with pytest.raises(TypeError):
            TranslationProvider()  # type: ignore[abstract]


class TestCreateProvider:
    """工厂函数根据配置创建正确的 Provider。"""

    def test_creates_openai_provider_for_openai_config(self):
        """TRANSLATE_API_PROVIDER=openai 时返回 OpenAICompatibleProvider"""
        from translator import create_translation_provider, OpenAICompatibleProvider

        provider = create_translation_provider("openai", api_key="sk-test")
        assert isinstance(provider, OpenAICompatibleProvider)

    def test_creates_openai_provider_for_deepseek_config(self):
        """TRANSLATE_API_PROVIDER=deepseek 时返回 OpenAICompatibleProvider（兼容协议）"""
        from translator import create_translation_provider, OpenAICompatibleProvider

        provider = create_translation_provider("deepseek", api_key="sk-test")
        assert isinstance(provider, OpenAICompatibleProvider)

    def test_creates_openai_provider_for_custom_config(self):
        """TRANSLATE_API_PROVIDER=custom 时返回 OpenAICompatibleProvider（兼容协议）"""
        from translator import create_translation_provider, OpenAICompatibleProvider

        provider = create_translation_provider("custom", api_key="sk-test")
        assert isinstance(provider, OpenAICompatibleProvider)

    def test_creates_claude_provider_for_anthropic_config(self):
        """TRANSLATE_API_PROVIDER=anthropic 时返回 ClaudeProvider"""
        from translator import create_translation_provider, ClaudeProvider

        provider = create_translation_provider("anthropic", api_key="sk-test")
        assert isinstance(provider, ClaudeProvider)

    def test_raises_on_unknown_provider(self):
        """未知 provider 名称时抛出 ValueError"""
        from translator import create_translation_provider

        with pytest.raises(ValueError, match="unknown|不支持|未知"):
            create_translation_provider("unknown_provider", api_key="sk-test")

    def test_raises_on_empty_api_key(self):
        """API Key 为空时抛出明确错误，不崩溃"""
        from translator import create_translation_provider

        with pytest.raises(ValueError, match="[Aa][Pp][Ii].*[Kk]ey|API"):
            create_translation_provider("openai", api_key="")


# ============================================================
# 切片2: _load_system_prompt
# ============================================================

class TestLoadSystemPrompt:
    """加载 prompts/translation_persona.txt 作为 System Prompt。"""

    def test_returns_non_empty_string(self):
        """返回非空字符串"""
        from translator import _load_system_prompt

        prompt = _load_system_prompt()
        assert isinstance(prompt, str)
        assert len(prompt.strip()) > 0

    def test_contains_persona_content(self):
        """加载的内容包含角色定义（Wildberries 俄语电商文案优化师）"""
        from translator import _load_system_prompt

        prompt = _load_system_prompt()
        assert "Wildberries 俄语电商文案优化师" in prompt

    def test_contains_json_format_instruction(self):
        """加载的内容包含 JSON 输出格式指令"""
        from translator import _load_system_prompt

        prompt = _load_system_prompt()
        assert "russian_title" in prompt
        assert "core_keywords" in prompt
        assert "russian_description" in prompt

    def test_raises_if_file_missing(self, monkeypatch):
        """文件不存在时抛出明确错误"""
        from translator import _load_system_prompt

        # 临时替换路径为不存在的文件
        monkeypatch.setattr(
            "translator._PERSONA_PATH",
            os.path.join(os.path.dirname(__file__), "nonexistent.txt"),
        )
        with pytest.raises(FileNotFoundError):
            _load_system_prompt()


# ============================================================
# 切片3: _post_process_title 标题后处理
# ============================================================

class TestPostProcessTitle:
    """俄语标题后处理规则。"""

    def test_lowercases_title(self):
        """标题转为全小写"""
        from text_utils import post_process_title

        result = post_process_title("Масляный Насос для Авто")
        assert result == "масляный насос для авто"

    def test_truncates_to_60_chars(self):
        """超过 60 字符的标题被截断"""
        from text_utils import post_process_title

        long_title = "а" * 80  # 80 个俄语字符
        result = post_process_title(long_title)
        assert len(result) <= 60
        assert result == "а" * 60

    def test_keeps_title_under_60_unchanged_length(self):
        """60 字符以内的标题长度不变（除小写外）"""
        from text_utils import post_process_title

        title = "бутылка для масла стеклянная с распылителем"  # ~47 chars
        result = post_process_title(title)
        assert len(result) <= 60
        assert result == title.lower()

    def test_removes_special_characters(self):
        """去除特殊符号，保留字母、数字、空格"""
        from text_utils import post_process_title

        result = post_process_title("масляный насос!!! (премиум) #1")
        # 感叹号、括号、井号应被移除
        assert "!" not in result
        assert "(" not in result
        assert ")" not in result
        assert "#" not in result
        # 保留字母数字和空格
        assert "масляный" in result
        assert "насос" in result
        assert "премиум" in result
        assert "1" in result

    def test_strips_whitespace(self):
        """去除首尾空格"""
        from text_utils import post_process_title

        result = post_process_title("  масляный насос  ")
        assert result == "масляный насос"
        assert not result.startswith(" ")
        assert not result.endswith(" ")

    def test_handles_empty_title(self):
        """空标题返回空字符串，不崩溃"""
        from text_utils import post_process_title

        result = post_process_title("")
        assert result == ""

    def test_handles_only_special_chars(self):
        """只有特殊符号的标题返回空字符串"""
        from text_utils import post_process_title

        result = post_process_title("!!!@@@###")
        assert result == ""

    def test_preserves_cyrillic(self):
        """保留俄语西里尔字母"""
        from text_utils import post_process_title

        result = post_process_title("Бутылка для Масла")
        # 所有西里尔字母应被保留（转为小写）
        assert "бутылка" in result
        assert "для" in result
        assert "масла" in result


# ============================================================
# 切片4: _parse_ai_response — JSON 解析与重试
# ============================================================

VALID_AI_JSON = """
{
  "russian_title": "бутылка для масла стеклянная",
  "core_keywords": "бутылка для масла стеклянная банка",
  "russian_description": "Стеклянная бутылка для масла с распылителем."
}
"""

VALID_AI_JSON_INLINE = (
    '{"russian_title":"бутылка для масла","core_keywords":"бутылка",'
    '"russian_description":"Описание товара."}'
)

JSON_IN_MARKDOWN = """
```json
{
  "russian_title": "масляный насос",
  "core_keywords": "насос масляный автомобильный",
  "russian_description": "Автомобильный масляный насос высокого качества."
}
```
"""


class TestParseAiResponse:
    """从 AI 响应中解析 JSON。"""

    def test_parses_valid_json_object(self):
        """正确解析标准 JSON 对象"""
        from translator import _parse_ai_response

        result = _parse_ai_response(VALID_AI_JSON)
        assert result["russian_title"] == "бутылка для масла стеклянная"
        assert result["core_keywords"] == "бутылка для масла стеклянная банка"
        assert "russian_description" in result

    def test_parses_inline_json(self):
        """正确解析单行 JSON"""
        from translator import _parse_ai_response

        result = _parse_ai_response(VALID_AI_JSON_INLINE)
        assert result["russian_title"] == "бутылка для масла"

    def test_extracts_json_from_markdown_block(self):
        """从 Markdown 代码块中提取 JSON"""
        from translator import _parse_ai_response

        result = _parse_ai_response(JSON_IN_MARKDOWN)
        assert result["russian_title"] == "масляный насос"

    def test_raises_on_invalid_json(self):
        """无效 JSON 时抛出 ValueError"""
        from translator import _parse_ai_response

        with pytest.raises(ValueError, match="[Jj][Ss][Oo][Nn]|解析"):
            _parse_ai_response("这不是合法的 JSON 响应")

    def test_raises_on_empty_response(self):
        """空响应时抛出 ValueError"""
        from translator import _parse_ai_response

        with pytest.raises(ValueError, match="[Jj][Ss][Oo][Nn]|解析|空"):
            _parse_ai_response("")

    def test_raises_on_missing_required_fields(self):
        """缺少必要字段时抛出 ValueError"""
        from translator import _parse_ai_response

        with pytest.raises(ValueError, match="缺少|字段|russian_title|core_keywords|russian_description"):
            _parse_ai_response('{"russian_title": "только заголовок"}')


class TestProviderRetriesOnParseFailure:
    """AI JSON 解析失败时自动重试一次。"""

    def test_retries_once_on_json_parse_error(self):
        """第一次返回无效 JSON → 重试 → 第二次成功"""
        from translator import OpenAICompatibleProvider

        call_count = [0]

        class RetryMockProvider(OpenAICompatibleProvider):
            def _call_api(self, title: str, details: str) -> str:
                call_count[0] += 1
                if call_count[0] == 1:
                    return "无效的 JSON 响应!!!"
                else:
                    return VALID_AI_JSON_INLINE

        provider = RetryMockProvider(api_key="sk-test")
        result = provider.translate("Test Product", "Test details")
        assert call_count[0] == 2
        assert result["russian_title"] == "бутылка для масла"

    def test_raises_after_two_failures(self):
        """两次解析都失败时抛出错误"""
        from translator import OpenAICompatibleProvider

        class AlwaysFailProvider(OpenAICompatibleProvider):
            def _call_api(self, title: str, details: str) -> str:
                return "总是返回无效 JSON"

        provider = AlwaysFailProvider(api_key="sk-test")
        with pytest.raises(ValueError, match="[Jj][Ss][Oo][Nn]|解析|重试"):
            provider.translate("Test", "Test details")


# ============================================================
# 切片6: translate_batch — 批量翻译主逻辑
# ============================================================

SAMPLE_PRODUCTS = [
    {
        "asin": "B0GVYXC124",
        "图片url": "https://img1.jpg",
        "标题": "5-in-1 Face Sculpting Machine",
        "详情": "Multifunctional body contouring machine.",
    },
    {
        "asin": "B0F45N6NS7",
        "图片url": "https://img2.jpg",
        "标题": "Anti Cellulite Massager",
        "详情": "Handheld electric anti-cellulite device.",
    },
]

# 模拟 AI 返回的原始 JSON 字段（与 _parse_ai_response 期望一致）
MOCK_AI_RESPONSE = {
    "core_keywords": "массажер для лица",
    "russian_title": "массажер для лица 5 в 1",
    "russian_description": "Многофункциональный массажер для лица и тела.",
}


class TestTranslateBatch:
    """批量翻译端到端行为。"""

    def test_translates_multiple_products(self, monkeypatch):
        """为多个产品批量翻译，返回结果包含原始+翻译字段"""
        from translator import translate_batch, OpenAICompatibleProvider

        class MockProvider(OpenAICompatibleProvider):
            def translate(self, title: str, details: str) -> dict:
                return dict(MOCK_AI_RESPONSE)

        def mock_create(*args, **kwargs):
            return MockProvider(api_key="sk-test")

        monkeypatch.setattr(
            "translator.create_translation_provider", mock_create
        )

        results = translate_batch(SAMPLE_PRODUCTS)
        assert len(results) == 2

        # 第 1 个产品：原始字段 + 翻译字段
        assert results[0]["asin"] == "B0GVYXC124"
        assert results[0]["标题"] == "5-in-1 Face Sculpting Machine"
        assert results[0]["核心流量词"] == "массажер для лица"
        assert results[0]["俄语标题"] == "массажер для лица 5 в 1"
        assert results[0]["俄语详情"] == "Многофункциональный массажер для лица и тела."

        # 第 2 个产品
        assert results[1]["asin"] == "B0F45N6NS7"
        assert results[1]["核心流量词"] == "массажер для лица"

    def test_progress_callback_fires_after_each_product(self, monkeypatch):
        """每次翻译完成后调用 progress_callback"""
        from translator import translate_batch, OpenAICompatibleProvider

        class MockProvider(OpenAICompatibleProvider):
            def translate(self, title: str, details: str) -> dict:
                return dict(MOCK_AI_RESPONSE)

        def mock_create(*args, **kwargs):
            return MockProvider(api_key="sk-test")

        monkeypatch.setattr(
            "translator.create_translation_provider", mock_create
        )

        progress = []
        results = translate_batch(
            SAMPLE_PRODUCTS,
            progress_callback=lambda i, total: progress.append((i, total)),
        )
        assert len(progress) == 2
        # 第一次回调: (1, 2)
        assert progress[0] == (1, 2)
        # 第二次回调: (2, 2)
        assert progress[1] == (2, 2)

    def test_empty_list_returns_empty_list(self, monkeypatch):
        """空产品列表返回空列表"""
        from translator import translate_batch

        results = translate_batch([])
        assert results == []

    def test_no_api_key_returns_clear_error_not_crash(self, monkeypatch):
        """不配置 API Key 时 translate_batch 返回带 error 字段的结果，不崩溃"""
        from translator import translate_batch

        # 模拟工厂函数抛出 API Key 未配置错误
        def mock_create_raise(*args, **kwargs):
            raise ValueError("API Key 未配置")

        monkeypatch.setattr(
            "translator.create_translation_provider", mock_create_raise
        )

        results = translate_batch(SAMPLE_PRODUCTS)
        # 应返回结果列表（不会崩溃），但包含错误信息
        assert len(results) == 2
        for r in results:
            assert "error" in r, f"未配置 API Key 时应返回 error 字段: {r}"
            assert r["asin"] in ["B0GVYXC124", "B0F45N6NS7"]

    def test_preserves_empty_source_columns(self, monkeypatch):
        """输出中货源、采购价、商品类别为空"""
        from translator import translate_batch, OpenAICompatibleProvider

        class MockProvider(OpenAICompatibleProvider):
            def translate(self, title: str, details: str) -> dict:
                return dict(MOCK_AI_RESPONSE)

        monkeypatch.setattr(
            "translator.create_translation_provider",
            lambda *a, **kw: MockProvider(api_key="sk-test"),
        )

        results = translate_batch(SAMPLE_PRODUCTS[:1])
        # translate_batch 输出的字典应包含空列占位
        assert "货源" in results[0]
        assert results[0]["货源"] == ""
        assert results[0]["采购价"] == ""
        assert results[0]["商品类别"] == ""


# ============================================================
# 切片7: CLI 入口 — python translator.py --input ... --output ...
# ============================================================

class TestCliEntry:
    """命令行入口行为。"""

    def test_required_args_parsed_correctly(self):
        """--input 和 --output 参数被正确解析"""
        from translator import _parse_cli_args

        args = _parse_cli_args(["--input", "input.xlsx", "--output", "output.xlsx"])
        assert args.input == "input.xlsx"
        assert args.output == "output.xlsx"

    def test_missing_input_raises_system_exit(self):
        """缺少 --input 参数时 argparse 触发 SystemExit"""
        from translator import _parse_cli_args

        with pytest.raises(SystemExit):
            _parse_cli_args(["--output", "output.xlsx"])

    def test_missing_output_raises_system_exit(self):
        """缺少 --output 参数时 argparse 触发 SystemExit"""
        from translator import _parse_cli_args

        with pytest.raises(SystemExit):
            _parse_cli_args(["--input", "input.xlsx"])

    def test_cli_end_to_end(self, monkeypatch, tmp_path):
        """端到端 CLI：读输入 Excel → 翻译 → 写输出 Excel"""
        from translator import main as translator_main

        # 创建输入 Excel
        input_path = tmp_path / "input.xlsx"
        _create_excel(
            str(input_path),
            headers=["asin", "图片url", "标题", "详情"],
            rows=[
                ["B0GVYXC124", "https://img.jpg", "Test Product", "Test description."],
            ],
        )

        # Mock create_translation_provider 返回 mock provider
        class MockProvider:
            def translate(self, title, details):
                return {
                    "core_keywords": "тест ключевые слова",
                    "russian_title": "тестовый продукт",
                    "russian_description": "Тестовое описание продукта.",
                }

        monkeypatch.setattr(
            "translator.create_translation_provider",
            lambda *a, **kw: MockProvider(),
        )

        output_path = tmp_path / "output.xlsx"

        # 模拟 CLI args
        import sys
        original_argv = sys.argv
        try:
            sys.argv = [
                "translator.py",
                "--input", str(input_path),
                "--output", str(output_path),
            ]
            translator_main()
        finally:
            sys.argv = original_argv

        # 验证输出文件存在
        assert output_path.is_file(), f"输出文件未生成: {output_path}"

        # 验证输出 Excel 有 12 列
        import openpyxl
        wb = openpyxl.load_workbook(str(output_path), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        wb.close()
        assert len(headers) == 12, f"输出应为 12 列，实际: {len(headers)}"
        assert headers[4] == "核心流量词"
        assert headers[5] == "俄语标题"


# ============================================================
# 辅助函数
# ============================================================

def _create_excel(path: str, headers: list[str], rows: list[list]) -> None:
    """创建测试用 Excel 文件。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
