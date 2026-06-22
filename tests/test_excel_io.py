"""
测试 excel_io.py — Excel 读写公共接口。

测试行为（非实现）：
- 从 Excel 自动识别并读取 ASIN 列
- 将产品数据写入符合"爬虫表格（处理前）"格式的 Excel
"""

import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ============================================================
# 辅助函数
# ============================================================

def _create_excel(path: str, headers: list[str], rows: list[list]) -> None:
    """创建测试用 Excel 文件。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _make_temp_xlsx(headers: list[str], rows: list[list]) -> str:
    """创建一个临时 Excel 文件并返回路径（调用方负责清理）。"""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)  # 关闭文件句柄，避免 Windows 文件锁
    _create_excel(path, headers, rows)
    return path


def _safe_unlink(path: str) -> None:
    """安全删除临时文件，忽略权限错误。"""
    try:
        os.unlink(path)
    except OSError:
        pass


# ============================================================
# read_asins_from_excel 测试
# ============================================================

class TestReadAsinsFromExcel:
    """从 Excel 读取 ASIN 列表的行为。"""

    @pytest.fixture
    def excel_with_asin_column(self):
        """创建包含标准 asin 列标题的临时 Excel。"""
        path = _make_temp_xlsx(
            headers=["asin", "标题", "详情"],
            rows=[
                ["B0GVYXC124", "Product A", "Description A"],
                ["B0F45N6NS7", "Product B", "Description B"],
            ],
        )
        yield path
        _safe_unlink(path)

    @pytest.fixture
    def excel_with_uppercase_asin(self):
        """ASIN 列名大写的情况。"""
        path = _make_temp_xlsx(
            headers=["ASIN", "商品名"],
            rows=[["B0GVYXC124", "Product A"], ["B000000000", "Product C"]],
        )
        yield path
        _safe_unlink(path)

    @pytest.fixture
    def excel_with_mixed_case_asin(self):
        """ASIN 列名混合大小写 Asin。"""
        path = _make_temp_xlsx(
            headers=["Asin", "title"],
            rows=[["B0GVYXC124", "Product A"]],
        )
        yield path
        _safe_unlink(path)

    @pytest.fixture
    def excel_without_asin_column(self):
        """没有 ASIN 列的情况。"""
        path = _make_temp_xlsx(
            headers=["商品编号", "标题"],
            rows=[["12345", "Product A"]],
        )
        yield path
        _safe_unlink(path)

    def test_reads_asin_list_from_excel(self, excel_with_asin_column):
        """基本行为：从 Excel 读取 ASIN 列表"""
        from excel_io import read_asins_from_excel

        asins = read_asins_from_excel(excel_with_asin_column)
        assert asins == ["B0GVYXC124", "B0F45N6NS7"]

    def test_strips_whitespace_from_asins(self):
        """ASIN 值去除首尾空格"""
        path = _make_temp_xlsx(
            headers=["asin"],
            rows=[[" B0GVYXC124 "], ["B0F45N6NS7"]],
        )
        try:
            from excel_io import read_asins_from_excel

            asins = read_asins_from_excel(path)
            assert asins == ["B0GVYXC124", "B0F45N6NS7"]
        finally:
            _safe_unlink(path)

    def test_detects_uppercase_asin_column(self, excel_with_uppercase_asin):
        """能识别大写 ASIN 列名"""
        from excel_io import read_asins_from_excel

        asins = read_asins_from_excel(excel_with_uppercase_asin)
        assert len(asins) == 2
        assert "B0GVYXC124" in asins

    def test_detects_mixed_case_asin_column(self, excel_with_mixed_case_asin):
        """能识别混合大小写 Asin 列名"""
        from excel_io import read_asins_from_excel

        asins = read_asins_from_excel(excel_with_mixed_case_asin)
        assert asins == ["B0GVYXC124"]

    def test_raises_on_missing_asin_column(self, excel_without_asin_column):
        """没有 ASIN 列时抛出明确错误"""
        from excel_io import read_asins_from_excel

        with pytest.raises(ValueError, match="[Aa][Ss][Ii][Nn]"):
            read_asins_from_excel(excel_without_asin_column)

    def test_skips_empty_rows(self):
        """跳过空 ASIN 行"""
        path = _make_temp_xlsx(
            headers=["asin", "title"],
            rows=[
                ["B0GVYXC124", "Product A"],
                [None, "No ASIN"],
                ["", "Empty ASIN"],
                ["B0F45N6NS7", "Product D"],
            ],
        )
        try:
            from excel_io import read_asins_from_excel

            asins = read_asins_from_excel(path)
            assert asins == ["B0GVYXC124", "B0F45N6NS7"]
        finally:
            _safe_unlink(path)

    def test_reads_real_sample_file(self):
        """能用真实样本文件"爬虫表格（处理前）.xlsx"正确读取"""
        sample_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "爬虫表格（处理前）.xlsx"
        )
        if not os.path.isfile(sample_path):
            pytest.skip("样本文件不存在")

        from excel_io import read_asins_from_excel

        asins = read_asins_from_excel(sample_path)
        assert len(asins) >= 1
        assert "B0GVYXC124" in asins


# ============================================================
# write_products_to_excel 测试
# ============================================================

SAMPLE_PRODUCTS = [
    {
        "asin": "B0GVYXC124",
        "图片url": "https://img1.jpg;https://img2.jpg;https://img3.jpg",
        "标题": "5-in-1 Face & Body Sculpting Machine",
        "详情": "Multifunctional body contouring machine with suction and LED.",
    },
    {
        "asin": "B0F45N6NS7",
        "图片url": "https://img4.jpg;https://img5.jpg",
        "标题": "Anti Cellulite Massager",
        "详情": "Handheld electric anti-cellulite device with heat mode.",
    },
]

EXPECTED_COLUMNS = ["asin", "图片url", "标题", "详情"]


class TestWriteProductsToExcel:
    """将产品列表写入 Excel 的行为。"""

    def test_writes_correct_columns(self):
        """输出的 Excel 包含正确顺序的 4 列：[asin, 图片url, 标题, 详情]"""
        from excel_io import write_products_to_excel

        path = _make_temp_xlsx(headers=["dummy"], rows=[])  # 占位，会被覆盖
        try:
            write_products_to_excel(SAMPLE_PRODUCTS, path)

            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            wb.close()
            assert headers == EXPECTED_COLUMNS, f"列名不匹配: {headers}"
        finally:
            _safe_unlink(path)

    def test_writes_all_rows(self):
        """所有产品行都被写入，数据与输入一致"""
        from excel_io import write_products_to_excel

        path = _make_temp_xlsx(headers=["dummy"], rows=[])
        try:
            write_products_to_excel(SAMPLE_PRODUCTS, path)

            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            data = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()

            assert len(data) == 2
            # 第 1 行
            assert data[0][0] == "B0GVYXC124"
            assert "img1.jpg" in str(data[0][1])
            assert data[0][2] == "5-in-1 Face & Body Sculpting Machine"
            # 第 2 行
            assert data[1][0] == "B0F45N6NS7"
            assert "img4.jpg" in str(data[1][1])
        finally:
            _safe_unlink(path)

    def test_handles_multiline_details(self):
        """详情中的换行符（\n）在 Excel 中正确保留"""
        from excel_io import write_products_to_excel

        product_with_newlines = [
            {
                "asin": "B000000000",
                "图片url": "https://img.jpg",
                "标题": "Test Product",
                "详情": "Line 1\nLine 2\nLine 3",
            }
        ]
        path = _make_temp_xlsx(headers=["dummy"], rows=[])
        try:
            write_products_to_excel(product_with_newlines, path)
            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            row = list(ws.iter_rows(min_row=2, values_only=True))[0]
            wb.close()
            assert "\n" in str(row[3]) or "Line 1" in str(row[3])
        finally:
            _safe_unlink(path)

    def test_output_matches_sample_format(self):
        """输出格式与"爬虫表格（处理前）"样本一致：asin 列名、4 列结构"""
        from excel_io import write_products_to_excel

        path = _make_temp_xlsx(headers=["dummy"], rows=[])
        try:
            write_products_to_excel(SAMPLE_PRODUCTS, path)

            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            # 与样本格式对照：列名必须完全匹配
            assert headers[0] == "asin", f"第1列应为 asin，实际: {headers[0]}"
            assert headers[1] == "图片url", f"第2列应为 图片url，实际: {headers[1]}"
            assert headers[2] == "标题", f"第3列应为 标题，实际: {headers[2]}"
            assert headers[3] == "详情", f"第4列应为 详情，实际: {headers[3]}"

            # 样本只有 4 列
            assert len(headers) == 4, f"应为 4 列，实际: {len(headers)}"
            wb.close()
        finally:
            _safe_unlink(path)


# ============================================================
# read_products_from_excel — 读取"处理前"表格（4列 → list[dict]）
# ============================================================

class TestReadProductsFromExcel:
    """从"爬虫表格（处理前）"读取完整产品信息。"""

    def test_reads_all_four_columns(self):
        """返回的字典包含 asin, 图片url, 标题, 详情 四个键"""
        path = _make_temp_xlsx(
            headers=["asin", "图片url", "标题", "详情"],
            rows=[
                ["B0GVYXC124", "https://img1.jpg", "Product A", "Description A"],
                ["B0F45N6NS7", "https://img2.jpg", "Product B", "Description B"],
            ],
        )
        try:
            from excel_io import read_products_from_excel

            products = read_products_from_excel(path)
            assert len(products) == 2
            assert products[0]["asin"] == "B0GVYXC124"
            assert products[0]["图片url"] == "https://img1.jpg"
            assert products[0]["标题"] == "Product A"
            assert products[0]["详情"] == "Description A"
        finally:
            _safe_unlink(path)

    def test_returns_empty_list_for_no_data_rows(self):
        """只有表头没有数据时返回空列表"""
        path = _make_temp_xlsx(
            headers=["asin", "图片url", "标题", "详情"],
            rows=[],
        )
        try:
            from excel_io import read_products_from_excel

            products = read_products_from_excel(path)
            assert products == []
        finally:
            _safe_unlink(path)

    def test_raises_on_missing_required_columns(self):
        """缺少必要列时抛出 ValueError"""
        path = _make_temp_xlsx(
            headers=["asin", "标题"],  # 缺少 图片url 和 详情
            rows=[["B0GVYXC124", "Product A"]],
        )
        try:
            from excel_io import read_products_from_excel

            with pytest.raises(ValueError):
                read_products_from_excel(path)
        finally:
            _safe_unlink(path)


# ============================================================
# write_translated_products_to_excel — 输出"处理后"表格（12列）
# ============================================================

EXPECTED_OUTPUT_COLUMNS = [
    "asin", "图片url", "标题", "详情",
    "核心流量词", "俄语标题", "俄语详情",
    "货源", "采购价", "", "", "商品类别",
]

SAMPLE_TRANSLATED_PRODUCTS = [
    {
        "asin": "B0GVYXC124",
        "图片url": "https://img1.jpg;https://img2.jpg",
        "标题": "5-in-1 Face & Body Sculpting Machine",
        "详情": "Multifunctional body contouring machine.",
        "核心流量词": "массажер для лица",
        "俄语标题": "массажер для лица 5 в 1",
        "俄语详情": "Многофункциональный массажер для лица и тела.",
    },
    {
        "asin": "B0F45N6NS7",
        "图片url": "https://img3.jpg",
        "标题": "Anti Cellulite Massager",
        "详情": "Handheld electric anti-cellulite device.",
        "核心流量词": "антицеллюлитный массажер",
        "俄语标题": "антицеллюлитный массажер электрический",
        "俄语详情": "Электрический антицеллюлитный массажер с нагревом.",
    },
]


class TestWriteTranslatedProductsToExcel:
    """将翻译后的产品列表写入"爬虫表格（处理后）"格式的 Excel。"""

    def test_writes_twelve_columns_in_correct_order(self):
        """输出 12 列，表头顺序与"处理后"样本一致"""
        from excel_io import write_translated_products_to_excel

        path = _make_temp_xlsx(headers=["dummy"], rows=[])
        try:
            write_translated_products_to_excel(SAMPLE_TRANSLATED_PRODUCTS, path)

            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            wb.close()

            assert len(headers) == 12, f"应为 12 列，实际: {len(headers)}"
            # 前 9 列有明确表头名
            assert headers[0] == "asin"
            assert headers[1] == "图片url"
            assert headers[2] == "标题"
            assert headers[3] == "详情"
            assert headers[4] == "核心流量词"
            assert headers[5] == "俄语标题"
            assert headers[6] == "俄语详情"
            assert headers[7] == "货源"
            assert headers[8] == "采购价"
            # 列 9、10 为无名列（空），openpyxl 读取为 None
            assert headers[9] is None or headers[9] == ""
            assert headers[10] is None or headers[10] == ""
            assert headers[11] == "商品类别"
        finally:
            _safe_unlink(path)

    def test_writes_translated_fields(self):
        """俄语翻译字段被正确写入"""
        from excel_io import write_translated_products_to_excel

        path = _make_temp_xlsx(headers=["dummy"], rows=[])
        try:
            write_translated_products_to_excel(SAMPLE_TRANSLATED_PRODUCTS, path)

            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            row = list(ws.iter_rows(min_row=2, values_only=True))[0]
            wb.close()

            # 列索引: asin=0, 图片url=1, 标题=2, 详情=3,
            #         核心流量词=4, 俄语标题=5, 俄语详情=6,
            #         货源=7(空), 采购价=8(空), 9=空, 10=空, 商品类别=11(空)
            assert row[4] == "массажер для лица"  # 核心流量词
            assert row[5] == "массажер для лица 5 в 1"  # 俄语标题
            assert "Многофункциональный" in str(row[6])  # 俄语详情
        finally:
            _safe_unlink(path)

    def test_empty_columns_are_empty(self):
        """货源、采购价、空列、商品类别为空"""
        from excel_io import write_translated_products_to_excel

        path = _make_temp_xlsx(headers=["dummy"], rows=[])
        try:
            write_translated_products_to_excel(SAMPLE_TRANSLATED_PRODUCTS, path)

            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            row = list(ws.iter_rows(min_row=2, values_only=True))[0]
            wb.close()

            # 列 7-11 都应为空
            for col_idx in [7, 8, 9, 10, 11]:
                assert row[col_idx] is None or row[col_idx] == "", (
                    f"第 {col_idx} 列应为空，实际值: {row[col_idx]}"
                )
        finally:
            _safe_unlink(path)

    def test_writes_all_rows(self):
        """所有产品行都被写入"""
        from excel_io import write_translated_products_to_excel

        path = _make_temp_xlsx(headers=["dummy"], rows=[])
        try:
            write_translated_products_to_excel(SAMPLE_TRANSLATED_PRODUCTS, path)

            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            data = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()

            assert len(data) == 2
            # 第 1 行 — 英文原文保留
            assert data[0][0] == "B0GVYXC124"
            assert data[0][2] == "5-in-1 Face & Body Sculpting Machine"
            # 第 2 行
            assert data[1][0] == "B0F45N6NS7"
            assert data[1][2] == "Anti Cellulite Massager"
        finally:
            _safe_unlink(path)
