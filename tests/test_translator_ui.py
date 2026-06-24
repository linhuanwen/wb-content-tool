"""
测试 translator_ui.py — Web 翻译功能区业务逻辑。

测试行为（非实现）：
- 文件上传校验（扩展名、必要列检测）
- 翻译执行（正常流程 + 无 API Key + 部分失败）
- 下载文件生成（12列格式）
"""

import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# ============================================================
# 测试辅助函数
# ============================================================

def _create_excel(path: str, headers: list[str], rows: list[list]) -> None:
    """创建测试用 Excel 文件。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _make_temp_xlsx(headers: list[str], rows: list[list]) -> str:
    """创建一个临时 Excel 文件并返回路径（调用方负责清理）。"""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    _create_excel(path, headers, rows)
    return path


def _make_temp_csv() -> str:
    """创建一个临时 .csv 文件并返回路径（调用方负责清理）。"""
    fd, path = tempfile.mkstemp(suffix=".csv")
    os.close(fd)
    with open(path, "w", encoding="utf-8") as f:
        f.write("dummy,data\n")
    return path


def _safe_unlink(path: str) -> None:
    """安全删除临时文件。"""
    try:
        os.unlink(path)
    except OSError:
        pass


# ============================================================
# 切片 1: validate_translation_upload — 文件扩展名校验
# ============================================================

class TestValidateTranslationUpload:
    """上传文件校验行为。"""

    def test_rejects_non_xlsx_file(self):
        """上传 .csv 文件时返回友好错误提示"""
        from translator_ui import validate_translation_upload

        csv_path = _make_temp_csv()
        try:
            result = validate_translation_upload(csv_path)
            assert result.is_valid is False
            assert ".xlsx" in result.error.lower() or "Excel" in result.error
            assert result.products == []
            assert result.count == 0
        finally:
            _safe_unlink(csv_path)

    # --- 切片 2：必要列校验 ---

    def test_rejects_excel_missing_required_columns(self):
        """上传有效 .xlsx 但缺少 [asin, 图片url, 标题, 详情] 时返回错误"""
        from translator_ui import validate_translation_upload

        path = _make_temp_xlsx(
            headers=["asin", "标题"],  # 缺少 图片url 和 详情
            rows=[["B0GVYXC124", "Product A"]],
        )
        try:
            result = validate_translation_upload(path)
            assert result.is_valid is False
            assert result.error  # 有错误信息
            assert result.products == []
            assert result.count == 0
        finally:
            _safe_unlink(path)

    # --- 切片 3：有效文件 → 返回产品列表 + 计数 ---

    def test_returns_products_and_count_for_valid_excel(self):
        """上传含全部必要列的有效 Excel → 返回产品列表和正确计数"""
        from translator_ui import validate_translation_upload

        path = _make_temp_xlsx(
            headers=["asin", "图片url", "标题", "详情"],
            rows=[
                ["B0GVYXC124", "https://img1.jpg", "Product A", "Description A"],
                ["B0F45N6NS7", "https://img2.jpg", "Product B", "Description B"],
                ["B000000000", "https://img3.jpg", "Product C", "Description C"],
            ],
        )
        try:
            result = validate_translation_upload(path)
            assert result.is_valid is True
            assert result.error == ""
            assert len(result.products) == 3
            assert result.count == 3
            assert result.products[0]["asin"] == "B0GVYXC124"
            assert result.products[1]["标题"] == "Product B"
        finally:
            _safe_unlink(path)


# ============================================================
# execute_translation 测试
# ============================================================

SAMPLE_PRODUCTS_FOR_TRANSLATION = [
    {
        "asin": "B0GVYXC124",
        "图片url": "https://img1.jpg",
        "标题": "5-in-1 Face Sculpting Machine",
        "详情": "Multifunctional body contouring machine.",
    },
    {
        "asin": "B0F45N6NS7",
        "图片url": "https://img2.jpg",
        "标题": "Anti Cellulite Massager",
        "详情": "Handheld electric anti-cellulite device.",
    },
]

MOCK_TRANSLATED_FIELDS = {
    "core_keywords": "массажер для лица",
    "russian_title": "массажер для лица 5 в 1",
    "russian_description": "Многофункциональный массажер для лица и тела.",
}


def _make_mock_provider():
    """创建一个返回固定翻译结果的 mock provider。"""
    from translator import TranslationProvider

    class MockProvider(TranslationProvider):
        def _call_api(self, title: str, details: str) -> str:
            return ""

        def translate(self, title: str, details: str) -> dict:
            return dict(MOCK_TRANSLATED_FIELDS)

    return MockProvider()


class TestExecuteTranslation:
    """翻译编排行为。"""

    # --- 切片 4：正常翻译流程 ---

    def test_translates_products_and_returns_translated(self):
        """正常翻译：所有产品成功 → translated 含翻译字段 + 空列"""
        from translator_ui import execute_translation

        result = execute_translation(
            SAMPLE_PRODUCTS_FOR_TRANSLATION,
            _provider_override=_make_mock_provider(),
        )

        assert len(result.translated) == 2
        assert result.failed == []

        # 第 1 个产品：原始字段 + 翻译字段 + 空列
        p0 = result.translated[0]
        assert p0["asin"] == "B0GVYXC124"
        assert p0["标题"] == "5-in-1 Face Sculpting Machine"
        assert p0["核心流量词"] == "массажер для лица"
        assert p0["俄语标题"] == "массажер для лица 5 в 1"
        assert p0["俄语详情"] == "Многофункциональный массажер для лица и тела."
        assert p0["货源"] == ""
        assert p0["采购价"] == ""
        assert p0["商品类别"] == ""

        # 第 2 个产品
        p1 = result.translated[1]
        assert p1["asin"] == "B0F45N6NS7"
        assert p1["核心流量词"] == "массажер для лица"

    # --- 切片 5：无 API Key 不崩溃 ---

    def test_no_api_key_returns_all_failed_not_crash(self, monkeypatch):
        """不配置 API Key 时所有产品入 failed，不抛异常"""
        from translator_ui import execute_translation

        # 通过 monkeypatch 让 create_translation_provider 抛出 API Key 错误
        import translator
        original = translator.create_translation_provider

        def mock_create_raise(*args, **kwargs):
            raise ValueError("API Key 未配置。请在 .env 文件中设置 TRANSLATE_API_KEY。")

        monkeypatch.setattr(translator, "create_translation_provider", mock_create_raise)

        result = execute_translation(SAMPLE_PRODUCTS_FOR_TRANSLATION)

        # 不应崩溃，全部入 failed
        assert result.translated == []
        assert len(result.failed) == 2
        for item in result.failed:
            assert "error" in item
            assert "API Key" in item["error"]
            assert item["asin"] in ["B0GVYXC124", "B0F45N6NS7"]

    # --- 切片 6：部分翻译失败 ---

    def test_partial_failure_mixes_translated_and_failed(self):
        """部分产品翻译成功 → translated + failed 同时有数据"""
        from translator_ui import execute_translation
        from translator import TranslationProvider

        class FlakyProvider(TranslationProvider):
            def _call_api(self, title: str, details: str) -> str:
                return ""

            def translate(self, title: str, details: str) -> dict:
                if "Anti" in title:
                    raise RuntimeError("翻译服务暂时不可用")
                return dict(MOCK_TRANSLATED_FIELDS)

        result = execute_translation(
            SAMPLE_PRODUCTS_FOR_TRANSLATION,
            _provider_override=FlakyProvider(),
        )

        # 第 1 个产品（不含 "Anti"）应成功
        assert len(result.translated) == 1
        assert result.translated[0]["asin"] == "B0GVYXC124"

        # 第 2 个产品（含 "Anti"）应失败
        assert len(result.failed) == 1
        assert result.failed[0]["asin"] == "B0F45N6NS7"
        assert "翻译服务暂时不可用" in result.failed[0]["error"]

    # --- 切片 7：进度回调 ---

    def test_progress_callback_fires_after_each_product(self):
        """每个产品翻译完成后触发 progress_callback(current, total)"""
        from translator_ui import execute_translation

        progress = []

        result = execute_translation(
            SAMPLE_PRODUCTS_FOR_TRANSLATION,
            _provider_override=_make_mock_provider(),
            progress_callback=lambda i, total: progress.append((i, total)),
        )

        assert len(progress) == 2
        assert progress[0] == (1, 2)
        assert progress[1] == (2, 2)


# ============================================================
# generate_translation_download 测试
# ============================================================

SAMPLE_TRANSLATED_PRODUCTS = [
    {
        "asin": "B0GVYXC124",
        "图片url": "https://img1.jpg",
        "标题": "5-in-1 Face Sculpting Machine",
        "详情": "Multifunctional body contouring machine.",
        "核心流量词": "массажер для лица",
        "俄语标题": "массажер для лица 5 в 1",
        "俄语详情": "Многофункциональный массажер для лица.",
        "货源": "",
        "采购价": "",
        "商品类别": "",
    },
    {
        "asin": "B0F45N6NS7",
        "图片url": "https://img2.jpg",
        "标题": "Anti Cellulite Massager",
        "详情": "Handheld electric anti-cellulite device.",
        "核心流量词": "антицеллюлитный массажер",
        "俄语标题": "антицеллюлитный массажер",
        "俄语详情": "Электрический антицеллюлитный массажер.",
        "货源": "",
        "采购价": "",
        "商品类别": "",
    },
]


class TestGenerateTranslationDownload:
    """下载文件生成行为。"""

    def test_returns_valid_xlsx_bytes_with_12_columns(self):
        """返回非空 xlsx 字节流，12 列，格式与"处理后"一致"""
        from translator_ui import generate_translation_download

        data = generate_translation_download(SAMPLE_TRANSLATED_PRODUCTS)

        assert isinstance(data, bytes)
        assert len(data) > 0

        # 验证可以被 openpyxl 重新打开
        import openpyxl
        import io

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        wb.close()

        assert len(headers) == 12, f"应为 12 列，实际: {len(headers)}"
        assert headers[0] == "asin"
        assert headers[1] == "图片url"
        assert headers[2] == "标题"
        assert headers[3] == "详情"
        assert headers[4] == "核心流量词"
        assert headers[5] == "俄语标题"
        assert headers[6] == "俄语详情"
        assert headers[7] == "货源"
        assert headers[8] == "采购价"
        assert headers[11] == "商品类别"

    def test_rows_match_input_products(self):
        """下载的 xlsx 数据行与翻译后产品一致"""
        from translator_ui import generate_translation_download

        data = generate_translation_download(SAMPLE_TRANSLATED_PRODUCTS)

        import openpyxl
        import io

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        assert len(rows) == 2
        assert rows[0][0] == "B0GVYXC124"
        assert rows[0][4] == "массажер для лица"
        assert rows[1][0] == "B0F45N6NS7"
        assert rows[1][5] == "антицеллюлитный массажер"


# ============================================================
# 两阶段工作流测试：execute_phase1_extraction
# ============================================================

SAMPLE_HTML_CONTENT = "<html><body><h1>5-in-1 Face Sculpting Machine</h1><p>Beauty device</p></body></html>"


def _setup_test_html_dir(tmp_path, asin_html_map: dict[str, str]) -> str:
    """创建测试用 HTML 目录，返回路径。"""
    html_dir = tmp_path / "html"
    html_dir.mkdir()
    for asin, content in asin_html_map.items():
        (html_dir / f"{asin}.html").write_text(content, encoding="utf-8")
    return str(html_dir)


def _setup_test_db(tmp_path) -> str:
    """创建测试用 SQLite 数据库，返回路径。"""
    import sqlite3
    from db import init_db

    db_path = str(tmp_path / "test_products.db")
    db = sqlite3.connect(db_path)
    init_db(db)
    db.close()
    return db_path


class TestExecutePhase1Extraction:
    """Phase 1 信息萃取编排行为。"""

    def test_extracts_html_and_saves_to_db(self, tmp_path):
        """正常流程：读取 HTML → AI 萃取 → 写入 products 表"""
        from translator_ui import execute_phase1_extraction

        html_dir = _setup_test_html_dir(tmp_path, {
            "B0TEST001": SAMPLE_HTML_CONTENT,
            "B0TEST002": SAMPLE_HTML_CONTENT,
        })
        db_path = _setup_test_db(tmp_path)

        results = execute_phase1_extraction(
            asins=["B0TEST001", "B0TEST002"],
            html_dir=html_dir,
            db_path=db_path,
            _extractor_override="mock",
        )

        assert len(results) == 2
        assert all("error" not in r for r in results)
        assert results[0]["asin"] == "B0TEST001"

        # 验证数据库中有记录
        import sqlite3
        from db import get_product

        db = sqlite3.connect(db_path)
        product = get_product(db, "B0TEST001")
        db.close()
        assert product is not None
        assert product["asin"] == "B0TEST001"
        assert product["category"] == "Beauty & Personal Care"
        assert "Facial massage" in product["features"]

    def test_skips_missing_html_files(self, tmp_path):
        """HTML 文件不存在时返回 error，不崩溃"""
        from translator_ui import execute_phase1_extraction

        html_dir = _setup_test_html_dir(tmp_path, {
            "B0TEST001": SAMPLE_HTML_CONTENT,
        })
        db_path = _setup_test_db(tmp_path)

        results = execute_phase1_extraction(
            asins=["B0TEST001", "B0MISSING"],
            html_dir=html_dir,
            db_path=db_path,
            _extractor_override="mock",
        )

        assert len(results) == 2
        # 第一个成功
        assert "error" not in results[0]
        # 第二个失败（HTML 不存在）
        assert "error" in results[1]
        assert "HTML" in results[1]["error"]

    def test_progress_callback_fires_after_each_asin(self, tmp_path):
        """每处理一个 ASIN 触发 progress_callback(current, total)"""
        from translator_ui import execute_phase1_extraction

        html_dir = _setup_test_html_dir(tmp_path, {
            "B0TEST001": SAMPLE_HTML_CONTENT,
            "B0TEST002": SAMPLE_HTML_CONTENT,
        })
        db_path = _setup_test_db(tmp_path)

        progress = []
        results = execute_phase1_extraction(
            asins=["B0TEST001", "B0TEST002"],
            html_dir=html_dir,
            db_path=db_path,
            _extractor_override="mock",
            progress_callback=lambda i, total: progress.append((i, total)),
        )

        assert len(progress) == 2
        assert progress[0] == (1, 2)
        assert progress[1] == (2, 2)

    def test_handles_empty_asin_list(self, tmp_path):
        """空 ASIN 列表返回空结果，不崩溃"""
        from translator_ui import execute_phase1_extraction

        html_dir = _setup_test_html_dir(tmp_path, {})
        db_path = _setup_test_db(tmp_path)

        results = execute_phase1_extraction(
            asins=[],
            html_dir=html_dir,
            db_path=db_path,
            _extractor_override="mock",
        )

        assert results == []


# ============================================================
# 两阶段工作流测试：execute_phase2_generation
# ============================================================

def _seed_products_table(db_path: str, products: list[dict]) -> None:
    """向 products 表插入测试数据。"""
    import sqlite3
    from db import upsert_product

    db = sqlite3.connect(db_path)
    for p in products:
        upsert_product(db, p)
    db.close()


class TestExecutePhase2Generation:
    """Phase 2 文案生成编排行为。"""

    def test_generates_from_db_and_saves_translations(self, tmp_path):
        """正常流程：从 products 读取 → AI 生成 → 写入 translations 表"""
        from translator_ui import execute_phase2_generation

        db_path = _setup_test_db(tmp_path)
        _seed_products_table(db_path, [
            {"asin": "B0TEST001", "title": "Test Product", "category": "Beauty"},
            {"asin": "B0TEST002", "title": "Test Product 2", "category": "Kitchen"},
        ])

        results = execute_phase2_generation(
            asins=["B0TEST001", "B0TEST002"],
            db_path=db_path,
            _generator_override="mock",
        )

        assert len(results) == 2
        assert all("error" not in r for r in results)

        # 验证数据库中有翻译记录
        import sqlite3
        from db import get_translation

        db = sqlite3.connect(db_path)
        t1 = get_translation(db, "B0TEST001")
        db.close()
        assert t1 is not None
        assert t1["russian_title"]
        assert t1["core_keywords"]
        assert t1["russian_description"]

    def test_skips_products_not_in_db(self, tmp_path):
        """数据库中不存在的 ASIN 返回 error"""
        from translator_ui import execute_phase2_generation

        db_path = _setup_test_db(tmp_path)
        _seed_products_table(db_path, [
            {"asin": "B0TEST001", "title": "Test Product", "category": "Beauty"},
        ])

        results = execute_phase2_generation(
            asins=["B0TEST001", "B0MISSING"],
            db_path=db_path,
            _generator_override="mock",
        )

        assert len(results) == 2
        assert "error" not in results[0]
        assert "error" in results[1]
        assert "不存在" in results[1]["error"]

    def test_progress_callback_fires_after_each_asin(self, tmp_path):
        """每处理一个 ASIN 触发 progress_callback(current, total)"""
        from translator_ui import execute_phase2_generation

        db_path = _setup_test_db(tmp_path)
        _seed_products_table(db_path, [
            {"asin": "B0TEST001", "title": "Test Product", "category": "Beauty"},
        ])

        progress = []
        execute_phase2_generation(
            asins=["B0TEST001"],
            db_path=db_path,
            _generator_override="mock",
            progress_callback=lambda i, total: progress.append((i, total)),
        )

        assert progress == [(1, 1)]

    def test_handles_empty_asin_list(self, tmp_path):
        """空 ASIN 列表返回空结果"""
        from translator_ui import execute_phase2_generation

        db_path = _setup_test_db(tmp_path)
        results = execute_phase2_generation(
            asins=[],
            db_path=db_path,
            _generator_override="mock",
        )

        assert results == []


# ============================================================
# 两阶段工作流测试：generate_phase2_download
# ============================================================

class TestGeneratePhase2Download:
    """Phase 2 下载文件生成行为。"""

    def test_returns_12_column_xlsx_from_db(self, tmp_path):
        """从数据库读取 translations + products 生成 12 列 xlsx"""
        from translator_ui import generate_phase2_download

        db_path = _setup_test_db(tmp_path)
        _seed_products_table(db_path, [
            {
                "asin": "B0TEST001",
                "title": "Test Product",
                "details": "Test details",
                "image_urls": "https://img1.jpg",
            },
        ])

        # 同时插入翻译记录
        import sqlite3
        from db import upsert_translation

        db = sqlite3.connect(db_path)
        upsert_translation(db, {
            "asin": "B0TEST001",
            "russian_title": "тестовый продукт",
            "core_keywords": "тест ключевые слова",
            "russian_description": "Описание тестового продукта.",
            "phase1_model": "mock",
            "phase2_model": "mock",
        })
        db.close()

        data = generate_phase2_download(db_path=db_path, asins=["B0TEST001"])

        assert isinstance(data, bytes)
        assert len(data) > 0

        import openpyxl
        import io

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        wb.close()

        assert len(headers) == 12
        assert headers[0] == "asin"
        assert headers[4] == "核心流量词"
        assert headers[5] == "俄语标题"

    def test_handles_empty_asins(self, tmp_path):
        """空 ASIN 列表返回空 xlsx（仅有表头）"""
        from translator_ui import generate_phase2_download

        db_path = _setup_test_db(tmp_path)
        data = generate_phase2_download(db_path=db_path, asins=[])

        assert isinstance(data, bytes)
        assert len(data) > 0

        import openpyxl
        import io

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        assert len(headers) == 12
        assert len(rows) == 0  # 无数据行
